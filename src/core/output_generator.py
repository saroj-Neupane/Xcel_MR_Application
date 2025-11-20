import logging
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from .utils import Utils


class OutputGenerator:
    """Handles Excel output generation"""
    
    def __init__(self, config, mapping_data=None, qc_reader=None):
        self.config = config
        self.mapping_data = mapping_data or []
        self.qc_reader = qc_reader
        self._processed_data_cache = {}
    
    def write_output(self, result_data, output_file):
        """Write processed data to Excel output file"""
        try:
            # Filter out empty or invalid data
            # For template-only mode, allow rows with only Pole (To Pole can be empty/N/A)
            # For normal mode, require both Pole and To Pole
            original_count = len(result_data)
            filtered_data = []
            filtered_out_count = 0
            
            for row in result_data:
                # Check if row exists and has valid Pole data
                pole_val = row.get('Pole', '') if row else ''
                to_pole_val = row.get('To Pole', '') if row else ''
                
                # Check if this is a template-only row (has _excel_row and may have PDF data)
                is_template_only = '_excel_row' in row
                has_pdf_data = any(row.get(key) for key in ['Structure Type', 'Existing Load', 'Proposed Load'])
                
                # Validate: Pole must exist, To Pole required only if not template-only or if row has connection data
                pole_valid = pole_val and str(pole_val).strip() != ''
                
                if is_template_only or has_pdf_data:
                    # Template-only mode: Pole is required, To Pole is optional
                    if row and pole_valid:
                        filtered_data.append(row)
                    else:
                        filtered_out_count += 1
                        logging.debug(f"Filtered out template-only row: Pole='{pole_val}', To Pole='{to_pole_val}'")
                else:
                    # Normal mode: Both Pole and To Pole required
                    to_pole_valid = to_pole_val and str(to_pole_val).strip() != ''
                    if row and pole_valid and to_pole_valid:
                        filtered_data.append(row)
                    else:
                        filtered_out_count += 1
                        logging.debug(f"Filtered out row: Pole='{pole_val}', To Pole='{to_pole_val}'")
            
            if filtered_out_count > 0:
                logging.debug(f"Filtered out {filtered_out_count} invalid rows")
            
            if not filtered_data:
                logging.warning("No valid data to write")
                return

            # Sort data using shared utility function
            sorted_data = sorted(filtered_data, key=lambda x: Utils.extract_numeric_part(x.get('Pole', '')))

            # Create data cache for QC sheet population
            # Clear previous cache to prevent stale data
            self._processed_data_cache.clear()
            for row in sorted_data:
                pole = row.get('Pole', '').strip()
                if pole:
                    self._processed_data_cache[pole] = row

            # Get template file path from config
            template_file = self.config.get('template_file', 'TEST_FILES_XCEL/Xcel Template.xlsx')
            template_path = Path(template_file)
            
            # Validate the template file before loading
            if not template_path.exists() or template_path.stat().st_size == 0:
                logging.error(f"Template file '{template_file}' is missing or empty.")
                return

            # Attempt to load the workbook from template inside a try/except block to catch EOFError
            # Always load fresh to avoid stale template data (template may have been modified externally)
            try:
                # Use keep_vba=True only for .xlsm files, not for .xlsx files
                if template_path.suffix.lower() == '.xlsm':
                    wb = load_workbook(template_file, keep_vba=True)
                else:
                    wb = load_workbook(template_file)
            except EOFError as eof_error:
                logging.error(f"EOFError encountered when loading template workbook '{template_file}': {eof_error}. The template file may be corrupted.")
                return
            except Exception as e:
                logging.error(f"Error loading template workbook '{template_file}': {e}")
                return

            # Determine worksheet to use
            if hasattr(self, 'config') and self.config:
                worksheet_name = self.config.get('output_settings', {}).get('worksheet_name', 'Consumers pg1')
            else:
                worksheet_name = 'Consumers pg1'

            if worksheet_name in wb.sheetnames:
                ws = wb[worksheet_name]
            else:
                ws = wb.active
                logging.warning(f"Worksheet '{worksheet_name}' not found, using '{ws.title}'")

            # Write data; using mapped writing if available, else a simple write
            if hasattr(self, 'mapping_data') and self.mapping_data:
                self._write_data_to_worksheet(ws, sorted_data, self.mapping_data)
            else:
                self._write_data_simple(ws, sorted_data)

            # Automatically populate QC sheet if QC reader is active
            if self.qc_reader and self.qc_reader.is_active():
                logging.info("Populating QC sheet")
                self._populate_qc_sheet(wb)
                self._add_sheet_comparison_formatting(wb, worksheet_name)

            wb.save(output_file)
            logging.info(f"Wrote {len(sorted_data)} records")

        except Exception as e:
            logging.error(f"Error writing output: {e}", exc_info=True)
            raise
    
    def _write_data_to_worksheet(self, ws, sorted_data, mapping_data):
        """Write data to worksheet using column mappings"""
        # Get output settings
        output_settings = self.config.get('output_settings', {})
        header_row = output_settings.get('header_row', 2)
        data_start_row = output_settings.get('data_start_row', 3)
        
        # Create mapping from output column names to Excel column indices
        column_mapping = {}
        for element, attribute, output_column in mapping_data:
            if output_column not in column_mapping:
                # Find the column in the header row
                for col_idx in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=header_row, column=col_idx)
                    if header_cell.value and str(header_cell.value).strip() == output_column:
                        column_mapping[output_column] = col_idx
                        break
        
        # Write data rows - preserve Pole and ToPole from template, populate other fields
        for row_idx, row_data in enumerate(sorted_data):
            # Use Excel row from template if available, otherwise use sequential
            excel_row = row_data.get('_excel_row', data_start_row + row_idx)
            
            # Set line number
            if 'Line No.' in column_mapping:
                ws.cell(row=excel_row, column=column_mapping['Line No.']).value = row_idx + 1
            
            # Write mapped data - skip Pole and ToPole columns to preserve template data
            for element, attribute, output_column in mapping_data:
                if output_column in column_mapping:
                    # Skip Pole and ToPole columns - preserve template data, but allow Pole Tag
                    if output_column.lower() in ['pole', 'to pole']:
                        continue
                        
                    col_idx = column_mapping[output_column]
                    
                    # Get the internal key for this mapping
                    internal_key = self._get_internal_key(element, attribute)
                    value = row_data.get(internal_key, "")
                    ws.cell(row=excel_row, column=col_idx).value = value
    
    def _write_data_simple(self, ws, sorted_data):
        """Simple data writing without mappings"""
        # Find the first row with data to determine column structure
        if not sorted_data:
            return
        
        # Get output settings
        output_settings = self.config.get('output_settings', {})
        data_start_row = output_settings.get('data_start_row', 2)
        
        # Use the keys from the first row as column headers
        headers = list(sorted_data[0].keys())
        
        # Write headers (assuming row 2) - skip Pole and ToPole to preserve template
        for col_idx, header in enumerate(headers, 1):
            if header.lower() not in ['pole', 'to pole']:
                ws.cell(row=2, column=col_idx).value = header
        
        # Write data (starting from data_start_row) - skip Pole and ToPole to preserve template
        for row_idx, row_data in enumerate(sorted_data):
            excel_row = data_start_row + row_idx
            
            for col_idx, header in enumerate(headers, 1):
                if header.lower() not in ['pole', 'to pole']:
                    value = row_data.get(header, "")
                    ws.cell(row=excel_row, column=col_idx).value = value
    
    def _get_internal_key(self, element, attribute):
        """Get the internal key used in row data for a given element/attribute mapping"""
        # Handle special cases
        if element == "Pole" and attribute == "SCID":
            return "Pole"
        elif element == "Pole" and attribute == "Number":
            return "Pole"
        elif element == "Pole" and attribute == "Tag":
            return "Pole Tag"
        elif element == "Pole" and attribute == "Latitude":
            return "Pole Latitude"
        elif element == "Pole" and attribute == "Longitude":
            return "Pole Longitude"
        elif element == "Pole" and attribute == "To Pole":
            return "To Pole"
        elif element == "Pole" and attribute == "Line No.":
            return "Line No."
        elif element == "Pole" and attribute == "Span Distance":
            return "Span Length"
        elif element == "Pole" and attribute == "Height & Class":
            return "Pole Height & Class"
        elif element == "Pole" and attribute == "Pole Height/Class":
            return "Pole Height & Class"
        elif element == "Pole" and attribute == "Address":
            return "Pole Address"
        elif element == "Pole" and attribute == "Guy Info":
            return "Guy Direction"  # Or could be "Guy Lead" depending on mapping
        elif element == "Pole" and attribute == "Existing Risers":
            return "Existing Risers"
        elif element == "Pole" and attribute == "Number of Existing Risers":
            return "Existing Risers"
        elif element == "Pole" and attribute == "MR Notes":
            return "MR Notes"
        elif element == "Power" and attribute == "Height":
            return "Power Height"
        elif element == "Power" and attribute == "Lowest Height":
            return "Power Height"
        elif element == "Power" and attribute == "Midspan":
            return "Power Midspan"
        elif element == "Streetlight" and attribute == "Height":
            return "Streetlight (bottom of bracket)"
        elif element == "Street Light" and attribute == "Height":
            return "Street Light Height"
        elif element == "Street Light" and attribute == "Lowest Height":
            return "Street Light Height"
        elif element in ["comm1", "comm2", "comm3", "comm4"] and attribute == "Height":
            return element
        elif element in ["comm1", "comm2", "comm3", "comm4"] and attribute == "Attachment Ht":
            return element
        elif element in ["comm1", "comm2", "comm3", "comm4"] and attribute == "Midspan Ht":
            return element
        elif element == "All_Comm_Heights" and attribute == "Summary":
            return "All Communication Heights"
        elif element == "Total_Comm_Count" and attribute == "Count":
            return "Total Communication Count"
        elif element in self.config.get("telecom_providers", []) and attribute == "Attachment Ht":
            return element
        elif element in self.config.get("telecom_providers", []) and attribute == "Midspan Ht":
            return element
        elif element == "Power Equipment" and attribute == "Equipment List":
            return "Power Equipments"
        elif element == "Pole" and attribute == "Existing Structure Type":
            return "Structure Type"
        elif element == "Pole" and attribute == "Existing Loading":
            return "Existing Load"
        elif element == "Pole" and attribute == "Proposed Loading":
            return "Proposed Load"
        elif element == "New Guy" and attribute == "Required":
            return "Guy Needed"
        else:
            # Default: use element name as key
            return element
    
    def _populate_qc_sheet(self, workbook):
        """Populate QC sheet with data from QC file"""
        # This is a complex method that would be moved from PoleDataProcessor
        # For now, we'll keep a simplified version
        try:
            if "QC" not in workbook.sheetnames:
                return
            
            # QC sheet population implementation
            # (would be moved from PoleDataProcessor)
            
        except Exception as e:
            logging.error(f"Error populating QC sheet: {e}")
    
    def _add_sheet_comparison_formatting(self, workbook, main_sheet_name):
        """Add conditional formatting to compare sheets"""
        try:
            # Sheet comparison formatting implementation
            # (would be moved from PoleDataProcessor)
            
        except Exception as e:
            logging.error(f"Error adding sheet comparison formatting: {e}")
    
    def generate_output_file(self, job_name, template_path):
        """Generate output file from template"""
        try:
            template_path = Path(template_path)
            if not template_path.exists():
                raise FileNotFoundError(f"Template file not found: {template_path}")
            
            # Create output directory in the same location as the template
            output_dir = template_path.parent / "output"
            output_dir.mkdir(exist_ok=True)
            
            # Create output filename (always .xlsx format)
            output_filename = f"{job_name}_MR_SS.xlsx"
            output_path = output_dir / output_filename
            
            # Copy template to output location
            import shutil
            shutil.copy2(template_path, output_path)
            return str(output_path)
            
        except Exception as e:
            logging.error(f"Error generating output file: {e}")
            raise 