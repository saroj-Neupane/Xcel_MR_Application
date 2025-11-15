import pandas as pd
import logging
from pathlib import Path
import re
import shutil
from openpyxl import load_workbook

from .utils import Utils
from .pdf_report_reader import PDFReportReader


class PoleDataProcessor:
    """Handles pole data processing and Excel output"""
    
    def __init__(self, config, geocoder=None, mapping_data=None, attachment_reader=None, qc_reader=None, pdf_reader=None, alden_qc_reader=None):
        self.config = config
        self.geocoder = geocoder
        self.mapping_data = mapping_data or []
        self.attachment_reader = attachment_reader
        self.qc_reader = qc_reader
        self.pdf_reader = pdf_reader
        self.alden_qc_reader = alden_qc_reader
        self.template_scids = None  # Will store SCIDs from template
        self.template_scids_by_sheet = {}  # Will store SCIDs from template organized by sheet
    
    def _is_end_marker(self, value):
        """Check if a value represents an END marker in the template."""
        if value is None:
            return False
        return str(value).strip().upper() == 'END'
    
    def _apply_end_marker(self, row_data):
        """Ensure END marker rows preserve END placeholders for span and midspan fields."""
        if not row_data:
            return row_data
        
        to_pole_value = row_data.get('To Pole', '')
        if not self._is_end_marker(to_pole_value):
            return row_data
        
        row_data['Span Length'] = 'END'
        
        for key in list(row_data.keys()):
            if 'Midspan' in key:
                row_data[key] = 'END'
        
        return row_data
    
    def read_template_scids(self, template_file_path):
        """Read SCIDs from Pole and To Pole columns in all sheets of template file"""
        try:
            import pandas as pd
            from pathlib import Path
            from openpyxl import load_workbook
            
            template_path = Path(template_file_path)
            if not template_path.exists():
                logging.error(f"Template file not found: {template_file_path}")
                return None
            
            # Load workbook to get all sheet names
            if template_path.suffix.lower() == '.xlsm':
                wb = load_workbook(template_file_path, keep_vba=True)
            else:
                wb = load_workbook(template_file_path)
            
            sheet_names = wb.sheetnames
            
            # Dictionary to store connections per sheet
            self.template_scids_by_sheet = {}
            
            # Read each sheet that has Pole and To Pole columns
            for sheet_name in sheet_names:
                try:
                    logging.debug(f"Reading worksheet '{sheet_name}'")
                    if template_path.suffix.lower() == '.xlsm':
                        df = pd.read_excel(template_file_path, sheet_name=sheet_name, keep_vba=True)
                    else:
                        df = pd.read_excel(template_file_path, sheet_name=sheet_name)
                    
                    # Look for Pole and To Pole columns with more flexible matching
                    pole_col = None
                    to_pole_col = None
                    
                    logging.debug(f"Available columns in '{sheet_name}': {list(df.columns)}")
                    
                    for col in df.columns:
                        col_str = str(col).strip().lower()
                        # Look for exact "Pole" column (not other columns containing "pole")
                        if col_str == 'pole':
                            pole_col = col
                            logging.debug(f"Found Pole column: '{col}'")
                        # Look for To Pole column
                        elif ('to pole' in col_str or 'topole' in col_str or col_str == 'to pole'):
                            to_pole_col = col
                            logging.debug(f"Found To Pole column: '{col}'")
                    
                    if pole_col is None or to_pole_col is None:
                        continue
                    
                    # Extract SCIDs from the columns with row information
                    template_connections = []
                    for idx, row in df.iterrows():
                        pole_scid = str(row[pole_col]).strip() if pd.notna(row[pole_col]) else ""
                        to_pole_scid = str(row[to_pole_col]).strip() if pd.notna(row[to_pole_col]) else ""
                        
                        # Include rows where Pole exists, even if To Pole is N/A or empty
                        if pole_scid and pole_scid != "nan":
                            # Store with Excel row number (pandas index + 2 for Excel row, accounting for header row)
                            excel_row = idx + 2
                            template_connections.append((pole_scid, to_pole_scid, excel_row))
                    
                    self.template_scids_by_sheet[sheet_name] = template_connections
                    
                except Exception as e:
                    continue
            
            # For backward compatibility, store the first sheet's connections in template_scids
            if self.template_scids_by_sheet:
                first_sheet = list(self.template_scids_by_sheet.keys())[0]
                self.template_scids = self.template_scids_by_sheet[first_sheet]
            else:
                self.template_scids = []
                logging.warning("No valid sheets with Pole/To Pole columns found in template")
            
            return self.template_scids
            
        except Exception as e:
            logging.error(f"Error reading template SCIDs: {e}")
            return None
    
    def _build_multi_sheet_template_rows(self, connections_df, mappings, sections_df):
        """Build rows based on all template SCIDs from all sheets"""
        temp_rows = {}
        
        # Collect all template connections from all sheets
        all_template_scids = []
        for sheet_name, sheet_scids in self.template_scids_by_sheet.items():
            all_template_scids.extend(sheet_scids)
        
        # Create connection span map to ensure consistent span distances
        connection_span_map = {}
        for row_idx, conn in connections_df.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            if n1 in mappings['node_id_to_scid'] and n2 in mappings['node_id_to_scid']:
                scid1 = mappings['node_id_to_scid'][n1]
                scid2 = mappings['node_id_to_scid'][n2]
                connection_key = tuple(sorted([scid1, scid2]))
                span_distance = conn.get('span_distance', '')
                
                if connection_key not in connection_span_map:
                    connection_span_map[connection_key] = span_distance
                    logging.debug(f"  Added to span map: {scid1}<->{scid2} = '{span_distance}' (row {row_idx})")
                else:
                    existing_span = connection_span_map[connection_key]
                    if existing_span != span_distance:
                        logging.warning(f"  CONFLICT: {scid1}<->{scid2} already has span '{existing_span}', ignoring new span '{span_distance}' (row {row_idx})")
                    else:
                        logging.debug(f"  Duplicate: {scid1}<->{scid2} = '{span_distance}' (row {row_idx})")
        
        for pole_scid, to_pole_scid, excel_row in all_template_scids:
            # Normalize SCIDs using the same logic as the main processing
            ignore_keywords = self.config.get('ignore_scid_keywords', [])
            pole_scid_norm = Utils.normalize_scid(pole_scid, ignore_keywords)
            to_pole_scid_norm = Utils.normalize_scid(to_pole_scid, ignore_keywords)
            
            # Find the connection data for this pair
            conn_info = self._find_connection_data(pole_scid_norm, to_pole_scid_norm, connections_df, mappings, connection_span_map)
            
            if conn_info:
                # Get pole node data
                pole_node_data = mappings['scid_to_node'].get(pole_scid_norm)
                if pole_node_data is not None:
                    # Create the output row
                    row_data = self._create_output_row(
                        pole_scid_norm, to_pole_scid_norm, conn_info, 
                        pole_node_data, mappings['scid_to_row'], sections_df
                    )
                    
                    if row_data:
                        # Validate that both Pole and To Pole SCIDs are valid before creating row
                        if not pole_scid or not to_pole_scid or pole_scid.strip() == '' or to_pole_scid.strip() == '':
                            logging.warning(f"Skipping template row creation: invalid Pole/To Pole values - Pole='{pole_scid}', To Pole='{to_pole_scid}'")
                        else:
                            # Use the original template SCIDs in the output
                            row_data['Pole'] = pole_scid
                            row_data['To Pole'] = to_pole_scid
                            # Store the Excel row number for precise positioning
                            row_data['_excel_row'] = excel_row
                            temp_rows[pole_scid_norm] = row_data
                            logging.debug(f"Created template-based row for {pole_scid} -> {to_pole_scid} at Excel row {excel_row}")
                    else:
                        logging.warning(f"Could not create row data for template connection {pole_scid} -> {to_pole_scid}")
                else:
                    logging.warning(f"Pole node data not found for template SCID: {pole_scid}")
            else:
                logging.warning(f"Connection data not found for template connection {pole_scid} -> {to_pole_scid}")
        
        logging.info(f"Created {len(temp_rows)} rows from {len(self.template_scids_by_sheet)} template sheets")
        return temp_rows
    
    def _build_template_based_rows(self, connections_df, mappings, sections_df):
        """Build rows based on template SCIDs instead of all connections"""
        temp_rows = {}
        
        logging.info(f"Processing {len(self.template_scids)} connections from template")
        
        # Create connection span map to ensure consistent span distances
        connection_span_map = {}
        logging.debug("Creating connection span map for template processing...")
        for row_idx, conn in connections_df.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            if n1 in mappings['node_id_to_scid'] and n2 in mappings['node_id_to_scid']:
                scid1 = mappings['node_id_to_scid'][n1]
                scid2 = mappings['node_id_to_scid'][n2]
                connection_key = tuple(sorted([scid1, scid2]))
                span_distance = conn.get('span_distance', '')
                
                if connection_key not in connection_span_map:
                    connection_span_map[connection_key] = span_distance
                    logging.debug(f"  Added to span map: {scid1}<->{scid2} = '{span_distance}' (row {row_idx})")
                else:
                    existing_span = connection_span_map[connection_key]
                    if existing_span != span_distance:
                        logging.warning(f"  CONFLICT: {scid1}<->{scid2} already has span '{existing_span}', ignoring new span '{span_distance}' (row {row_idx})")
                    else:
                        logging.debug(f"  Duplicate: {scid1}<->{scid2} = '{span_distance}' (row {row_idx})")
        
        for pole_scid, to_pole_scid, excel_row in self.template_scids:
            # Normalize SCIDs using the same logic as the main processing
            ignore_keywords = self.config.get('ignore_scid_keywords', [])
            pole_scid_norm = Utils.normalize_scid(pole_scid, ignore_keywords)
            to_pole_scid_norm = Utils.normalize_scid(to_pole_scid, ignore_keywords)
            
            # Find the connection data for this pair
            conn_info = self._find_connection_data(pole_scid_norm, to_pole_scid_norm, connections_df, mappings, connection_span_map)
            
            if conn_info:
                # Get pole node data
                pole_node_data = mappings['scid_to_node'].get(pole_scid_norm)
                if pole_node_data is not None:
                    # Create the output row
                    row_data = self._create_output_row(
                        pole_scid_norm, to_pole_scid_norm, conn_info, 
                        pole_node_data, mappings['scid_to_row'], sections_df
                    )
                    
                    if row_data:
                        # Validate that both Pole and To Pole SCIDs are valid before creating row
                        if not pole_scid or not to_pole_scid or pole_scid.strip() == '' or to_pole_scid.strip() == '':
                            logging.warning(f"Skipping template row creation: invalid Pole/To Pole values - Pole='{pole_scid}', To Pole='{to_pole_scid}'")
                        else:
                            # Use the original template SCIDs in the output
                            row_data['Pole'] = pole_scid
                            row_data['To Pole'] = to_pole_scid
                            # Store the Excel row number for precise positioning
                            row_data['_excel_row'] = excel_row
                            temp_rows[pole_scid_norm] = row_data
                            logging.debug(f"Created template-based row for {pole_scid} -> {to_pole_scid} at Excel row {excel_row}")
                    else:
                        logging.warning(f"Could not create row data for template connection {pole_scid} -> {to_pole_scid}")
                else:
                    logging.warning(f"Pole node data not found for template SCID: {pole_scid}")
            else:
                logging.warning(f"Connection data not found for template connection {pole_scid} -> {to_pole_scid}")
        
        logging.info(f"Created {len(temp_rows)} rows from template connections")
        return temp_rows
    
    def _process_multi_sheet_template_connections(self, connections_df, mappings, sections_df):
        """Process connections from all sheets in template"""
        result_data = []
        
        # Collect all template connections from all sheets
        all_template_scids = []
        for sheet_name, sheet_scids in self.template_scids_by_sheet.items():
            logging.info(f"Adding {len(sheet_scids)} connections from sheet '{sheet_name}'")
            all_template_scids.extend(sheet_scids)
        
        logging.info(f"Processing {len(all_template_scids)} total template connections from {len(self.template_scids_by_sheet)} sheets")
        
        for pole_scid, to_pole_scid, excel_row in all_template_scids:
            # Normalize SCIDs using the same logic as the main processing
            ignore_keywords = self.config.get('ignore_scid_keywords', [])
            pole_scid_norm = Utils.normalize_scid(pole_scid, ignore_keywords)
            to_pole_scid_norm = Utils.normalize_scid(to_pole_scid, ignore_keywords)
            
            logging.info(f"Processing template connection: {pole_scid} -> {to_pole_scid} (normalized: {pole_scid_norm} -> {to_pole_scid_norm})")
            
            # Check if To Pole is N/A or empty - treat as invalid connection
            is_invalid_connection = (
                not to_pole_scid or 
                self._is_end_marker(to_pole_scid) or
                to_pole_scid.upper() == "N/A" or 
                to_pole_scid == "nan" or 
                to_pole_scid.strip() == ""
            )
            
            # Find the connection data for this pair (only if not invalid)
            conn_info = None
            if not is_invalid_connection:
                conn_info = self._find_connection_data(pole_scid_norm, to_pole_scid_norm, connections_df, mappings)
            
            # Get pole node data
            pole_node_data = mappings['scid_to_node'].get(pole_scid_norm)
            if pole_node_data is not None:
                if conn_info:
                    # Create the output row with connection data
                    row_data = self._create_output_row(
                        pole_scid_norm, to_pole_scid_norm, conn_info, 
                        pole_node_data, mappings['scid_to_row'], sections_df
                    )
                else:
                    # Create the output row without connection data (pole-only data)
                    if is_invalid_connection:
                        logging.info(f"Invalid To Pole value '{to_pole_scid}' for {pole_scid}, creating pole-only row")
                    else:
                        logging.info(f"No valid connection found for {pole_scid} -> {to_pole_scid}, creating pole-only row")
                    row_data = self._create_pole_only_row(
                        pole_scid_norm, to_pole_scid_norm, 
                        pole_node_data, mappings['scid_to_row'], sections_df
                    )
                
                if row_data:
                    # Use the original template SCIDs in the output
                    row_data['Pole'] = pole_scid
                    row_data['To Pole'] = to_pole_scid
                    row_data = self._apply_end_marker(row_data)
                    # Store the Excel row number for precise positioning
                    row_data['_excel_row'] = excel_row
                    result_data.append(row_data)
                    logging.debug(f"Created template-based output row for {pole_scid} -> {to_pole_scid} at Excel row {excel_row}")
                else:
                    logging.warning(f"Could not create output row for template connection {pole_scid} -> {to_pole_scid}")
            else:
                logging.warning(f"Pole node data not found for template SCID: {pole_scid}")
        
        logging.info(f"Generated {len(result_data)} output rows from {len(self.template_scids_by_sheet)} template sheets")
        return result_data
    
    def _process_template_based_connections(self, connections_df, mappings, sections_df):
        """Process only the connections specified in the template"""
        result_data = []
        
        logging.info(f"Processing {len(self.template_scids)} template connections for output generation")
        
        for pole_scid, to_pole_scid, excel_row in self.template_scids:
            # Normalize SCIDs using the same logic as the main processing
            ignore_keywords = self.config.get('ignore_scid_keywords', [])
            pole_scid_norm = Utils.normalize_scid(pole_scid, ignore_keywords)
            to_pole_scid_norm = Utils.normalize_scid(to_pole_scid, ignore_keywords)
            
            logging.info(f"Processing template connection: {pole_scid} -> {to_pole_scid} (normalized: {pole_scid_norm} -> {to_pole_scid_norm})")
            
            # Check if To Pole is N/A or empty - treat as invalid connection
            is_invalid_connection = (
                not to_pole_scid or 
                self._is_end_marker(to_pole_scid) or
                to_pole_scid.upper() == "N/A" or 
                to_pole_scid == "nan" or 
                to_pole_scid.strip() == ""
            )
            
            # Find the connection data for this pair (only if not invalid)
            conn_info = None
            if not is_invalid_connection:
                conn_info = self._find_connection_data(pole_scid_norm, to_pole_scid_norm, connections_df, mappings)
            
            # Get pole node data
            pole_node_data = mappings['scid_to_node'].get(pole_scid_norm)
            if pole_node_data is not None:
                if conn_info:
                    # Create the output row with connection data
                    row_data = self._create_output_row(
                        pole_scid_norm, to_pole_scid_norm, conn_info, 
                        pole_node_data, mappings['scid_to_row'], sections_df
                    )
                else:
                    # Create the output row without connection data (pole-only data)
                    if is_invalid_connection:
                        logging.info(f"Invalid To Pole value '{to_pole_scid}' for {pole_scid}, creating pole-only row")
                    else:
                        logging.info(f"No valid connection found for {pole_scid} -> {to_pole_scid}, creating pole-only row")
                    row_data = self._create_pole_only_row(
                        pole_scid_norm, to_pole_scid_norm, 
                        pole_node_data, mappings['scid_to_row'], sections_df
                    )
                
                if row_data:
                    # Use the original template SCIDs in the output
                    row_data['Pole'] = pole_scid
                    row_data['To Pole'] = to_pole_scid
                    row_data = self._apply_end_marker(row_data)
                    # Store the Excel row number for precise positioning
                    row_data['_excel_row'] = excel_row
                    result_data.append(row_data)
                    logging.debug(f"Created template-based output row for {pole_scid} -> {to_pole_scid} at Excel row {excel_row}")
                else:
                    logging.warning(f"Could not create output row for template connection {pole_scid} -> {to_pole_scid}")
            else:
                logging.warning(f"Pole node data not found for template SCID: {pole_scid}")
        
        logging.info(f"Generated {len(result_data)} output rows from template connections")
        return result_data
    
    def _create_pole_only_row(self, pole_scid, to_pole_scid, pole_node_data, scid_to_row, sections_df):
        """Create a row with pole-specific data but no connection-specific data"""
        try:
            # Start with basic pole data
            row_data = {
                'Pole': pole_scid,
                'To Pole': to_pole_scid,
                'Pole Tag': pole_node_data.get('pole_tag_tagtext', ''),
                'Latitude': Utils.round_coordinate(pole_node_data.get('latitude', '')),
                'Longitude': Utils.round_coordinate(pole_node_data.get('longitude', '')),
                'Pole Height/Class': self._format_pole_height_class(pole_node_data),
                'Structure Type': Utils.clean_structure_type(pole_node_data.get('structure_type', '')),
                'Notes': pole_node_data.get('mr_note', ''),
            }
            
            # Process attachment data using the same logic as _create_output_row
            # Create empty section since this is pole-only (no connection-specific data)
            empty_section = pd.Series()
            mapped_elements = self._get_mapped_elements()
            
            # Process attachments for pole-only row (is_pole_to_reference=True to skip midspan data)
            attachment_result = self._process_attachments(pole_node_data, empty_section, mapped_elements, pole_scid, is_pole_to_reference=True)
            row_data.update(attachment_result)
            
            # Skip section data for pole-only rows (no connection-specific data)
            # Section data is only relevant for connections between poles
            
            # Add PDF report data (Structure Type, Existing Load, Proposed Load)
            if self.pdf_reader:
                try:
                    # Extract pole number from SCID (e.g., "001" -> 1)
                    pole_number = self._extract_pole_number_from_scid(pole_scid)
                    if pole_number:
                        pdf_data = self.pdf_reader.extract_pole_data(pole_number)
                        row_data['Structure Type'] = Utils.clean_structure_type(pdf_data.get('structure_type', ''))
                        row_data['Existing Load'] = pdf_data.get('existing_load', '')
                        row_data['Proposed Load'] = pdf_data.get('proposed_load', '')
                        logging.debug(f"Added PDF data for pole-only row {pole_scid} (number {pole_number}): {pdf_data}")
                    else:
                        logging.debug(f"Could not extract pole number from SCID {pole_scid} for pole-only row")
                except Exception as e:
                    logging.error(f"Error extracting PDF data for pole-only row {pole_scid}: {e}")
            
            # Add New Guy Required field based on MR Notes
            mr_notes = pole_node_data.get('mr_note', '')
            row_data['Guy Needed'] = self._determine_new_guy_required(mr_notes)
            
            # Add Existing Risers count
            row_data['Existing Risers'] = self._count_existing_risers(pole_node_data, pole_scid)
            
            # Connection-specific fields should remain empty
            # These include: Span Length, connection_id, span_distance, etc.
            
            logging.debug(f"Created pole-only row for {pole_scid} -> {to_pole_scid}")
            return row_data
            
        except Exception as e:
            logging.error(f"Error creating pole-only row for {pole_scid} -> {to_pole_scid}: {e}")
            return None
    
    def _find_connection_data(self, pole_scid, to_pole_scid, connections_df, mappings, connection_span_map=None):
        """Find connection data for a specific pole pair"""
        logging.debug(f"Looking for connection data: {pole_scid} <-> {to_pole_scid}")
        
        # Look for connection in both directions
        for row_idx, conn in connections_df.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            if n1 in mappings['node_id_to_scid'] and n2 in mappings['node_id_to_scid']:
                scid1 = mappings['node_id_to_scid'][n1]
                scid2 = mappings['node_id_to_scid'][n2]
                
                # Check if this connection matches what we're looking for
                if (scid1 == pole_scid and scid2 == to_pole_scid) or (scid1 == to_pole_scid and scid2 == pole_scid):
                    # Use corrected span distance if available, otherwise use original
                    connection_key = tuple(sorted([scid1, scid2]))
                    if connection_span_map and connection_key in connection_span_map:
                        span_distance = connection_span_map[connection_key]
                        logging.debug(f"Using span distance from connection_span_map: {connection_key} = '{span_distance}'")
                    else:
                        span_distance = conn.get('span_distance', '')
                        logging.debug(f"Using span distance from connection row: '{span_distance}'")
                    
                    conn_info = {
                        'connection_id': conn.get('connection_id', ''),
                        'span_distance': span_distance,
                        'node1_id': n1,
                        'node2_id': n2,
                        'original_scid1': scid1,
                        'original_scid2': scid2
                    }
                    logging.debug(f"✓ MATCH: {pole_scid} <-> {to_pole_scid} found in row {row_idx}: span_distance='{conn_info['span_distance']}', connection_id='{conn_info['connection_id']}', node_ids=({n1},{n2}), original_scids=({scid1},{scid2})")
                    return conn_info
                else:
                    # Log potential matches for debugging common node issues
                    if pole_scid in [scid1, scid2] or to_pole_scid in [scid1, scid2]:
                        logging.debug(f"  Partial match in row {row_idx}: {scid1} <-> {scid2} (looking for {pole_scid} <-> {to_pole_scid})")
        
        logging.debug(f"✗ NO MATCH: No connection data found for {pole_scid} <-> {to_pole_scid}")
        return None
    
    def process_data(self, nodes_df, connections_df, sections_df, progress_callback=None, 
                    manual_routes=None, clear_existing_routes=False):
        """Process pole data"""
        if progress_callback:
            progress_callback(40, "Filtering pole data...")
        
        # Handle empty nodes_df - template-only processing
        if nodes_df.empty:
            logging.info("No nodes data available - processing template-only with PDF data")
            if hasattr(self, 'template_scids_by_sheet') and self.template_scids_by_sheet:
                total_connections = sum(len(scids) for scids in self.template_scids_by_sheet.values())
                logging.info(f"Using multi-sheet template SCIDs for processing: {total_connections} total connections from {len(self.template_scids_by_sheet)} sheets")
                # Process template-based connections without nodes data
                result_data = self._process_template_only_connections(connections_df, sections_df)
                return result_data
            elif self.template_scids:
                logging.info(f"Using template SCIDs for processing: {len(self.template_scids)} connections")
                # Process template-based connections without nodes data
                result_data = self._process_template_only_connections(connections_df, sections_df)
                return result_data
            else:
                logging.warning("No template connections found and no nodes data - cannot process")
                return []
        
        # Single data source: No caching needed for alternative lookups
        
        # Normalize SCIDs and filter nodes
        nodes_df = nodes_df.copy()
        ignore_keywords = self.config.get('ignore_scid_keywords', [])
        nodes_df['scid'] = nodes_df['scid'].apply(lambda x: Utils.normalize_scid(x, ignore_keywords))
        nodes_df = nodes_df.drop_duplicates(subset='scid')
        
        # Sort nodes by SCID numerically
        nodes_df['sort_key'] = nodes_df['scid'].apply(Utils.extract_numeric_part)
        nodes_df = nodes_df.sort_values(by='sort_key')
        nodes_df = nodes_df.drop('sort_key', axis=1)
        
        # Filter valid SCIDs: node_type = 'pole' OR 'reference' AND pole_status != 'underground'
        filtered = Utils.filter_valid_nodes(nodes_df)
        
        if filtered.empty:
            logging.warning("No valid pole or reference data found in nodes")
            # Try template-only processing if available
            if hasattr(self, 'template_scids_by_sheet') and self.template_scids_by_sheet:
                logging.info("Falling back to multi-sheet template-only processing")
                result_data = self._process_template_only_connections(connections_df, sections_df)
                return result_data
            elif self.template_scids:
                logging.info("Falling back to template-only processing")
                result_data = self._process_template_only_connections(connections_df, sections_df)
                return result_data
            else:
                raise ValueError("No valid pole or reference data found")
        
        # Log the filtered data for debugging
        poles_count = len(filtered[filtered['node_type'].str.strip().str.lower().eq('pole')])
        references_count = len(filtered[filtered['node_type'].str.strip().str.lower().eq('reference')])
        logging.info(f"Found {poles_count} valid poles and {references_count} valid references")
        
        # Create mappings
        mappings = self._create_mappings(nodes_df, filtered)
        
        if progress_callback:
            if not progress_callback(50, "Building connections..."):
                return []  # Stop processing if requested
        
        # Build temp rows - use template SCIDs if available, otherwise use all connections
        # Note: For multi-sheet templates, temp_rows are not used in the main processing path
        if hasattr(self, 'template_scids_by_sheet') and self.template_scids_by_sheet:
            # Multi-sheet template - build rows from all sheets
            temp_rows = self._build_multi_sheet_template_rows(connections_df, mappings, sections_df)
        elif self.template_scids:
            temp_rows = self._build_template_based_rows(connections_df, mappings, sections_df)
        else:
            temp_rows = self._build_temp_rows(connections_df, mappings, manual_routes, clear_existing_routes)
        
        # If QC file is active, only keep poles mentioned in QC file
        if self.qc_reader and self.qc_reader.is_active():
            qc_scids = self.qc_reader.get_qc_scids()
            original_count = len(temp_rows)
            temp_rows = {scid: data for scid, data in temp_rows.items() if scid in qc_scids}
            logging.info(f"QC filtering: reduced from {original_count} to {len(temp_rows)} poles (only QC SCIDs)")
            
            if not temp_rows:
                logging.warning("No poles found after QC filtering - check that QC SCIDs match pole SCIDs in data")
        
        if progress_callback:
            if not progress_callback(70, "Processing connections..."):
                return []  # Stop processing if requested

        # Process connections to generate output rows (one row per connection involving a pole)
        result_data = []
        
        # If QC file is active, generate output based on QC connections only
        if self.qc_reader and self.qc_reader.is_active():
            logging.info("QC file is active - filtering output to QC connections only")
            result_data = self._process_qc_filtered_connections(
                connections_df, mappings, sections_df
            )
        elif hasattr(self, 'template_scids_by_sheet') and self.template_scids_by_sheet:
            # Use multi-sheet template-based processing - process all template connections from all sheets
            logging.info(f"Using multi-sheet template-based processing - processing {len(self.template_scids_by_sheet)} sheets")
            result_data = self._process_multi_sheet_template_connections(connections_df, mappings, sections_df)
        elif self.template_scids:
            # Use single-sheet template-based processing - only process connections from template
            logging.info("Using single-sheet template-based processing - processing only template connections")
            result_data = self._process_template_based_connections(connections_df, mappings, sections_df)
        else:
            # No template connections found - cannot process without template
            logging.error("No template connections found - cannot process without template file")
            return []
        
        # Filter results based on manual routes if specified
        if manual_routes:
            manual_scids = {scid for route in manual_routes for scid in route['poles']}
            logging.info(f"Filtering results to manual route SCIDs: {sorted(manual_scids)}")
            
            original_count = len(result_data)
            result_data = [row for row in result_data if row.get('Pole') in manual_scids]
            logging.info(f"Manual route filtering: reduced from {original_count} to {len(result_data)} rows")
        
        # Sort result by pole SCID (unless QC file is active, then preserve QC order)
        if self.qc_reader and self.qc_reader.is_active():
            # QC file is active - result_data is already in QC order, don't sort
            logging.info("Preserving QC file order for output")
        else:
            # No QC file - sort by pole SCID as usual
            result_data.sort(key=lambda x: Utils.extract_numeric_part(x.get('Pole', '')))
        
        if progress_callback:
            if not progress_callback(90, f"Generated {len(result_data)} output rows"):
                return []  # Stop processing if requested
        
        return result_data
    
    def _process_template_only_connections(self, connections_df, sections_df):
        """Process connections using only template SCIDs and PDF data"""
        logging.info("Processing template-only connections with PDF data")
        
        result_data = []
        
        # Check for multi-sheet template first
        template_scids_to_process = []
        if hasattr(self, 'template_scids_by_sheet') and self.template_scids_by_sheet:
            # Collect all template connections from all sheets
            for sheet_name, sheet_scids in self.template_scids_by_sheet.items():
                logging.info(f"Adding {len(sheet_scids)} connections from sheet '{sheet_name}' for template-only processing")
                template_scids_to_process.extend(sheet_scids)
            logging.info(f"Processing {len(template_scids_to_process)} total connections from {len(self.template_scids_by_sheet)} sheets")
        elif self.template_scids:
            template_scids_to_process = self.template_scids
            logging.info(f"Processing {len(template_scids_to_process)} connections from single template sheet")
        
        if not template_scids_to_process:
            logging.warning("No template SCIDs available for template-only processing")
            return result_data
        
        # Create empty mappings for template-only processing
        mappings = []
        
        for pole_scid, to_pole_scid, excel_row in template_scids_to_process:
            logging.debug(f"Processing template connection: {pole_scid} <-> {to_pole_scid}")
            
            # Create basic row data from template
            row_data = {
                'Pole': pole_scid,
                'To Pole': to_pole_scid,
                '_excel_row': excel_row,
                'Line No.': '',  # Will be set during output
            }
            
            # Try to find connection data if connections_df is available
            if not connections_df.empty:
                connection_data = self._find_connection_data(pole_scid, to_pole_scid, connections_df, mappings)
                if connection_data:
                    # Add connection-specific data
                    for key, value in connection_data.items():
                        if key not in ['Pole', 'To Pole']:  # Don't overwrite template values
                            row_data[key] = value
            
            # Add PDF data if available
            if self.pdf_reader:
                try:
                    pole_number = self._extract_pole_number_from_scid(pole_scid)
                    if pole_number:
                        pdf_data = self.pdf_reader.extract_pole_data(pole_number)
                        row_data['Structure Type'] = Utils.clean_structure_type(pdf_data.get('structure_type', ''))
                        row_data['Existing Load'] = pdf_data.get('existing_load', '')
                        row_data['Proposed Load'] = pdf_data.get('proposed_load', '')
                        logging.debug(f"Added PDF data for pole {pole_scid}: {pdf_data}")
                except Exception as e:
                    logging.error(f"Error extracting PDF data for pole {pole_scid}: {e}")
            
            # Add attachment data if available
            if self.attachment_reader:
                try:
                    # Get attachment data for the pole
                    attachment_data = self.attachment_reader.get_scid_data(pole_scid)
                    if not attachment_data.empty:
                        # Add basic attachment info (can be expanded based on needs)
                        row_data['Attachment Count'] = len(attachment_data)
                        logging.debug(f"Found {len(attachment_data)} attachments for pole {pole_scid}")
                except Exception as e:
                    logging.error(f"Error getting attachment data for pole {pole_scid}: {e}")
            
            row_data = self._apply_end_marker(row_data)
            result_data.append(row_data)
        
        logging.info(f"Template-only processing completed: {len(result_data)} rows")
        return result_data
    
    def _create_mappings(self, nodes_df, filtered):
        """Create various lookup mappings"""
        return {
            'node_id_to_scid': nodes_df.set_index('node_id')['scid'].to_dict(),
            'scid_to_row': nodes_df.set_index('scid').to_dict('index'),
            'scid_to_node': nodes_df.set_index('scid').to_dict('index'),  # Alias for scid_to_row
            'node_id_to_row': nodes_df.set_index('node_id').to_dict('index'),
            'valid_poles': set(filtered['node_id'])
        }
    
    def _process_standard_connections(self, connections_df, mappings, sections_df):
        """Process standard connections without QC filtering (optimized)"""
        result_data = []
        processed_connections = set()
        
        # Pre-filter connections to only valid poles for better performance
        valid_poles = mappings['valid_poles']
        mask = (connections_df['node_id_1'].isin(valid_poles)) & (connections_df['node_id_2'].isin(valid_poles))
        valid_connections = connections_df[mask]
        
        logging.info(f"Processing {len(valid_connections)} valid connections out of {len(connections_df)} total connections")
        
        # Check for duplicate connections in the data
        connection_span_map = {}
        duplicate_connections = []
        logging.debug("Creating connection span map for main processing...")
        for row_idx, conn in valid_connections.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            if n1 in mappings['node_id_to_scid'] and n2 in mappings['node_id_to_scid']:
                scid1 = mappings['node_id_to_scid'][n1]
                scid2 = mappings['node_id_to_scid'][n2]
                connection_key = tuple(sorted([scid1, scid2]))
                span_distance = conn.get('span_distance', '')
                connection_id = conn.get('connection_id', '')
                
                if connection_key in connection_span_map:
                    existing_span = connection_span_map[connection_key]
                    if existing_span != span_distance:
                        duplicate_connections.append(f"{scid1}<->{scid2}: existing='{existing_span}', new='{span_distance}'")
                        logging.warning(f"  CONFLICT in main processing: {scid1}<->{scid2} already has span '{existing_span}', new span '{span_distance}' (row {row_idx}, connection_id='{connection_id}')")
                else:
                    connection_span_map[connection_key] = span_distance
                    logging.debug(f"  Added to main span map: {scid1}<->{scid2} = '{span_distance}' (row {row_idx}, connection_id='{connection_id}')")
                    
                    # Special logging for connections involving poles 10, 11, 12 to help debug the issue
                    if any(pole in [scid1, scid2] for pole in ['10', '11', '12']):
                        logging.info(f"  SPAN MAP DEBUG: Connection {scid1}<->{scid2} = '{span_distance}' (row {row_idx}, connection_id='{connection_id}')")
        
        if duplicate_connections:
            logging.warning(f"Found {len(duplicate_connections)} connections with conflicting span lengths:")
            for dup in duplicate_connections[:10]:  # Show first 10 duplicates
                logging.warning(f"  {dup}")
            if len(duplicate_connections) > 10:
                logging.warning(f"  ... and {len(duplicate_connections) - 10} more")
        
        # Additional check for potential span length cross-contamination
        logging.info(f"Connection span map created with {len(connection_span_map)} unique connections")
        for connection_key, span_distance in list(connection_span_map.items())[:5]:  # Show first 5 for verification
            scid1, scid2 = connection_key
            logging.debug(f"Connection span map: {scid1}<->{scid2} = '{span_distance}'")
        
        # Check for potential SCID mapping issues
        scid_to_node_count = {}
        for node_id, scid in mappings['node_id_to_scid'].items():
            if scid in scid_to_node_count:
                scid_to_node_count[scid].append(node_id)
            else:
                scid_to_node_count[scid] = [node_id]
        
        duplicate_scids = {scid: nodes for scid, nodes in scid_to_node_count.items() if len(nodes) > 1}
        if duplicate_scids:
            logging.warning(f"Found {len(duplicate_scids)} SCIDs mapped to multiple node_ids:")
            for scid, nodes in list(duplicate_scids.items())[:5]:
                logging.warning(f"  SCID '{scid}' mapped to node_ids: {nodes}")
        
        for _, conn in valid_connections.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            
            # Use SCID-based connection key to avoid processing same connection twice
            scid1 = mappings['node_id_to_scid'][n1]
            scid2 = mappings['node_id_to_scid'][n2]
            connection_key = tuple(sorted([scid1, scid2]))
            
            if connection_key not in processed_connections:
                processed_connections.add(connection_key)
                
                # For duplicate connections, use the first occurrence (already in connection_span_map)
                span_distance = connection_span_map.get(connection_key, conn.get('span_distance', ''))
                
                node1_data = mappings['node_id_to_row'].get(n1, {})
                node2_data = mappings['node_id_to_row'].get(n2, {})
                node1_type = str(node1_data.get('node_type', '')).strip().lower()
                node2_type = str(node2_data.get('node_type', '')).strip().lower()
                
                conn_info = {
                    'connection_id': conn.get('connection_id', ''),
                    'span_distance': span_distance
                }
                
                # Debug logging for span length issues
                if conn_info.get('span_distance'):
                    logging.debug(f"Connection {scid1} <-> {scid2}: span_distance='{conn_info['span_distance']}', connection_id='{conn_info['connection_id']}'")
                    
                    # Special logging for connections involving poles 10, 11, 12 to help debug the issue
                    if any(pole in [scid1, scid2] for pole in ['10', '11', '12']):
                        logging.info(f"  CONNECTION DEBUG: Processing {scid1} <-> {scid2}: span_distance='{conn_info['span_distance']}', connection_id='{conn_info['connection_id']}', node_types=({node1_type},{node2_type})")
                
                # Validate SCIDs before processing
                if not scid1 or not scid2 or scid1.strip() == '' or scid2.strip() == '':
                    logging.warning(f"Skipping connection with invalid SCIDs: scid1='{scid1}', scid2='{scid2}'")
                    continue
                
                # Generate row(s) for this connection
                if node1_type == 'pole' and node2_type == 'reference':
                    # Pole -> Reference: Pole in "Pole" column, Reference in "To Pole" column
                    row_data = self._create_output_row(scid1, scid2, conn_info, node1_data, mappings['scid_to_row'], sections_df)
                    if row_data:
                        row_data = self._apply_end_marker(row_data)
                        result_data.append(row_data)
                        logging.debug(f"Added pole->reference connection: {scid1} -> {scid2}")
                elif node1_type == 'reference' and node2_type == 'pole':
                    # Reference -> Pole: Pole in "Pole" column, Reference in "To Pole" column
                    row_data = self._create_output_row(scid2, scid1, conn_info, node2_data, mappings['scid_to_row'], sections_df)
                    if row_data:
                        row_data = self._apply_end_marker(row_data)
                        result_data.append(row_data)
                        logging.debug(f"Added reference->pole connection: {scid2} -> {scid1}")
                elif node1_type == 'pole' and node2_type == 'pole':
                    # Pole -> Pole: First pole in "Pole" column, Second pole in "To Pole" column
                    row_data = self._create_output_row(scid1, scid2, conn_info, node1_data, mappings['scid_to_row'], sections_df)
                    if row_data:
                        row_data = self._apply_end_marker(row_data)
                        result_data.append(row_data)
                        logging.debug(f"Added pole->pole connection: {scid1} -> {scid2}")
        
        # Count pole-to-reference connections for logging
        pole_ref_count = sum(1 for row in result_data if row.get('To Pole', '') and 
                           any(ref_scid in row.get('To Pole', '') for ref_scid in 
                               [scid for scid, data in mappings['scid_to_row'].items() 
                                if str(data.get('node_type', '')).strip().lower() == 'reference']))
        
        logging.info(f"Generated {len(result_data)} total connections, including {pole_ref_count} pole-to-reference connections")
        
        return result_data
    
    def _build_temp_rows(self, connections_df, mappings, manual_routes, clear_existing_routes):
        """Build temporary rows for processing"""
        temp = {}
        processed = set()
        
        # Initialize all valid poles
        for node_id in mappings['valid_poles']:
            scid = mappings['node_id_to_scid'][node_id]
            node_data = mappings['node_id_to_row'].get(node_id, {})
            guy_info = self._extract_guy_info(node_data.get('mr_note', ''))
            
            temp[scid] = {
                'Pole': scid,
                'Guy Size': '',
                'Guy Lead': ', '.join(guy_info['leads']),
                'Guy Direction': ', '.join(guy_info['directions']),
                'To Pole': '',
                'connection_id': '',
                'span_distance': ''
            }
        
        # Skip Excel connection processing if QC file is active
        if self.qc_reader and self.qc_reader.is_active():
            logging.info("QC file is active - skipping Excel connection processing")
            connection_data = {}
        else:
            # Process Excel connections
            connection_data = self._process_excel_connections(
                connections_df, mappings, temp, processed, clear_existing_routes
            )
        
        # Apply manual routes (only if QC file is not active)
        if manual_routes and not (self.qc_reader and self.qc_reader.is_active()):
            self._apply_manual_routes(manual_routes, temp, connection_data)
        elif manual_routes and self.qc_reader and self.qc_reader.is_active():
            logging.info("QC file is active - manual routes will be ignored in favor of QC connections")
        
        logging.info(f"Built {len(temp)} pole records with routing information")
        return temp

    def _process_excel_connections(self, connections_df, mappings, temp, processed, clear_existing_routes):
        """Process connections from Excel data with enhanced reference node logic"""
        logging.info("Processing automatic connections from Excel data...")
        connection_data = {}
        
        for _, conn in connections_df.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            
            if (n1 in mappings['valid_poles'] and n2 in mappings['valid_poles']):
                connection_key = tuple(sorted([n1, n2]))
                if connection_key not in processed:
                    processed.add(connection_key)
                    scid1 = mappings['node_id_to_scid'][n1]
                    scid2 = mappings['node_id_to_scid'][n2]
                    
                    # Get node types to handle reference nodes correctly
                    node1_data = mappings['node_id_to_row'].get(n1, {})
                    node2_data = mappings['node_id_to_row'].get(n2, {})
                    node1_type = str(node1_data.get('node_type', '')).strip().lower()
                    node2_type = str(node2_data.get('node_type', '')).strip().lower()
                    
                    conn_info = {
                        'connection_id': conn.get('connection_id', ''),
                        'span_distance': conn.get('span_distance', '')
                    }
                    
                    # Store connection data (use sorted tuple as key to avoid duplication)
                    connection_key = tuple(sorted([scid1, scid2]))
                    connection_data[connection_key] = conn_info
                    
                    if not clear_existing_routes:
                        # Handle reference node logic: references must be at 'To Pole'
                        if node2_type == 'reference' and node1_type == 'pole':
                            # scid1 is pole, scid2 is reference
                            temp[scid1].update({'To Pole': scid2, **conn_info})
                        elif node1_type == 'reference' and node2_type == 'pole':
                            # scid2 is pole, scid1 is reference  
                            temp[scid2].update({'To Pole': scid1, **conn_info})
                        elif node1_type == 'pole' and node2_type == 'pole':
                            # Both are poles, use normal connection logic
                            temp[scid1].update({'To Pole': scid2, **conn_info})
                        else:
                            # Default behavior for other cases
                            temp[scid1].update({'To Pole': scid2, **conn_info})
        
        if clear_existing_routes:
            logging.info("Cleared existing route data as requested")
            for scid in temp:
                temp[scid]['To Pole'] = ''
        
        return connection_data

    def _apply_manual_routes(self, manual_routes, temp, connection_data):
        """Apply manual routes to pole data"""
        logging.info(f"Applying {len(manual_routes)} manual routes...")
        
        # Verify that every defined connection actually exists in the Excel data
        missing = []
        for route in manual_routes:
            for from_scid, to_scid in route['connections']:
                connection_key = tuple(sorted([from_scid, to_scid]))
                if from_scid not in temp or to_scid not in temp or connection_key not in connection_data:
                    missing.append((from_scid, to_scid))
        
        if missing:
            msg = "The following manual-route connections are invalid or missing:\n" + \
                  "\n".join(f"{a} → {b}" for a, b in missing)
            raise ValueError("Invalid manual routes detected")
        
        for route_idx, route in enumerate(manual_routes):
            logging.info(f"Processing manual route {route_idx + 1}: {' → '.join(route['poles'])}")
            for from_scid, to_scid in route['connections']:
                connection_key = tuple(sorted([from_scid, to_scid]))
                conn_info = connection_data.get(connection_key, {})
                
                temp[from_scid].update({
                    'To Pole': to_scid,
                    'connection_id': conn_info.get('connection_id', ''),
                    'span_distance': conn_info.get('span_distance', '')
                })
                logging.info(f"  Set {from_scid} → {to_scid}")
                
                if not conn_info.get('connection_id'):
                    logging.warning(f"  No Excel connection data found for {from_scid} → {to_scid}")
            
            # Handle dead-end
            last_pole = route['poles'][-1]
            if last_pole in temp:
                temp[last_pole]['To Pole'] = ''
                logging.info(f"  Set {last_pole} as dead-end")
    
    def _extract_guy_info(self, note):
        """Extract guy information from notes.
           Supports multiple formats:
             - "PL NEW SINGLE HELIX ANCHOR 15' S WITH OFFSET" -> Guy Lead = "15'", Guy Direction = "S"
             - "PL NEW xxxxxx ANCHOR 20'6" NW" -> Guy Lead = "20'6"", Guy Direction = "NW"
             - "ANCHOR 10' W"  -> Guy Lead = "10'" and Guy Direction = "W"
             - "ANCHOR 15'6" NW" -> Guy Lead = "15'6"" and Guy Direction = "NW"
             - "GUY 3/8" EHS 20' S" -> Guy Size = "3/8" EHS", Guy Lead = "20'", Guy Direction = "S"
             - "5/16" EHS GUY 15' N" -> Guy Size = "5/16" EHS", Guy Lead = "15'", Guy Direction = "N"
        """
        if not note or pd.isna(note):
            return {'leads': [], 'directions': [], 'sizes': []}
        
        note = str(note).upper()
        leads = []
        directions = []
        sizes = []
        
        # Pattern 0: PL NEW format - "PL NEW SINGLE HELIX ANCHOR 15' S WITH OFFSET"
        pl_new_pattern = r"PL\s+NEW\s+[A-Z\s]+\s+ANCHOR\s+(\d+)'(?:\s*(\d+)\")?\s+([NSEW]{1,2})(?:\s|$)"
        pl_new_matches = re.findall(pl_new_pattern, note)
        
        # If PL NEW patterns are found, only use those and skip other patterns
        if pl_new_matches:
            for feet, inches, direction in pl_new_matches:
                # Build Guy Lead string preserving inches if provided
                if inches:
                    lead = f"{feet}'{inches}\""
                else:
                    lead = f"{feet}'"
                lead = lead.strip()
                direction = direction.strip()
                combined = f"{lead} {direction}"
                if combined not in [f"{l} {d}" for l, d in zip(leads, directions)]:
                    leads.append(lead)
                    directions.append(direction)
                    sizes.append('')  # No size info in PL NEW format
        else:
            # Only process other patterns if no PL NEW patterns were found
            
            # Pattern 1: ANCHOR format - "ANCHOR 10' W"
            anchor_pattern = r"ANCHOR\s+(\d+)'(?:\s*(\d+)\")?\s+([NSEW]{1,2})"
            anchor_matches = re.findall(anchor_pattern, note)
            for feet, inches, direction in anchor_matches:
                # Build Guy Lead string preserving inches if provided
                if inches:
                    lead = f"{feet}'{inches}\""
                else:
                    lead = f"{feet}'"
                lead = lead.strip()
                direction = direction.strip()
                combined = f"{lead} {direction}"
                if combined not in [f"{l} {d}" for l, d in zip(leads, directions)]:
                    leads.append(lead)
                    directions.append(direction)
                    sizes.append('')  # No size info in ANCHOR format
            
            # Pattern 2: GUY with size - "GUY 3/8" EHS 20' S" or "5/16" EHS GUY 15' N"
            guy_pattern = r"(?:GUY\s+)?(\d+/\d+\"\s*EHS|[\d.]+\"\s*EHS)\s*(?:GUY\s+)?(\d+)'(?:\s*(\d+)\")?\s+([NSEW]{1,2})"
            guy_matches = re.findall(guy_pattern, note)
            for size, feet, inches, direction in guy_matches:
                # Build Guy Lead string preserving inches if provided
                if inches:
                    lead = f"{feet}'{inches}\""
                else:
                    lead = f"{feet}'"
                lead = lead.strip()
                direction = direction.strip()
                size = size.strip()
                combined = f"{lead} {direction}"
                if combined not in [f"{l} {d}" for l, d in zip(leads, directions)]:
                    leads.append(lead)
                    directions.append(direction)
                    sizes.append(size)
            
            # Pattern 3: General guy pattern - any remaining patterns with just lead/direction
            # Make this more restrictive to avoid matching height values
            general_pattern = r"(?:^|\s)(\d+)'(?:\s*(\d+)\")?\s+([NSEW]{1,2})(?:\s|$)"
            general_matches = re.findall(general_pattern, note)
            for feet, inches, direction in general_matches:
                if inches:
                    lead = f"{feet}'{inches}\""
                else:
                    lead = f"{feet}'"
                lead = lead.strip()
                direction = direction.strip()
                combined = f"{lead} {direction}"
                if combined not in [f"{l} {d}" for l, d in zip(leads, directions)]:
                    leads.append(lead)
                    directions.append(direction)
                    sizes.append('')  # No size info in general format
        
        return {'leads': leads, 'directions': directions, 'sizes': sizes}
    
    def _determine_new_guy_required(self, mr_notes):
        """Determine if new guy is required based on MR Notes.
        
        Args:
            mr_notes (str): The MR Notes text
            
        Returns:
            str: "YES" if "METRONET ANCHOR" or "METRONET DG" is found in MR Notes, "NO" otherwise
        """
        if not mr_notes or pd.isna(mr_notes):
            return "NO"
        
        mr_notes_str = str(mr_notes).upper().strip()
        
        # Check if either "METRONET ANCHOR" or "METRONET DG" is present
        if "METRONET ANCHOR" in mr_notes_str or "METRONET DG" in mr_notes_str:
            return "YES"
        
        return "NO"
    
    def _extract_pole_number_from_scid(self, scid):
        """Extract pole number from SCID for PDF file matching
        
        Args:
            scid: The SCID string (e.g., "001", "123", "001A")
            
        Returns:
            Integer pole number or None if extraction fails
        """
        try:
            # Remove any non-numeric characters and get the numeric part
            import re
            numeric_part = re.sub(r'[^0-9]', '', str(scid))
            if numeric_part:
                return int(numeric_part)
            return None
        except (ValueError, TypeError):
            logging.debug(f"Could not extract pole number from SCID: {scid}")
            return None
    
    def _is_telecom_company(self, company_name):
        """Check if a company is NOT the power company (i.e., it's a communication company)
        
        Args:
            company_name: The company name to check (should be lowercase)
            
        Returns:
            bool: True if the company is NOT the power company
        """
        try:
            if not company_name or not company_name.strip():
                return False
            
            # Get power company name from config
            power_company = self.config.get("power_company", "").strip().lower()
            if not power_company:
                # If no power company defined, assume all companies are telecom
                return True
            
            # Check if this is the power company
            return power_company not in company_name.lower()
            
        except Exception as e:
            logging.debug(f"Error checking if company is telecom provider: {company_name} - {e}")
            return False

    def _create_output_row(self, pole_scid, to_pole_scid, conn_info, pole_node_data, scid_to_row, sections_df):
        """Create an output row for a connection involving a pole"""
        try:
            # Validate that both Pole and To Pole SCIDs are valid
            if (not pole_scid or not to_pole_scid or 
                pole_scid.strip() == '' or to_pole_scid.strip() == '' or 
                self._is_end_marker(to_pole_scid)):
                logging.warning(f"Skipping row creation: invalid Pole/To Pole values - Pole='{pole_scid}', To Pole='{to_pole_scid}'")
                return None
            
            # Get pole data from the pole_scid (this should always be a pole, not a reference)
            node = scid_to_row.get(pole_scid, pole_node_data)
            
            # Determine connection type by checking if to_pole_scid is a reference
            to_pole_data = scid_to_row.get(to_pole_scid, {})
            to_pole_type = str(to_pole_data.get('node_type', '')).strip().lower()
            is_pole_to_reference = (to_pole_type == 'reference')
            
            logging.debug(f"Connection {pole_scid} -> {to_pole_scid}: to_pole_type='{to_pole_type}', is_pole_to_reference={is_pole_to_reference}")
            
            # Find section data for this connection
            connection_id = conn_info.get('connection_id', '')
            section = self._find_section(connection_id, sections_df, pole_scid, to_pole_scid)
            logging.debug(f"Section lookup for {pole_scid} -> {to_pole_scid}: connection_id='{connection_id}', section found: {section is not None}")
            
            # Single data source: Only use section data from 'sections' sheet via connection_id
            # No fallback mechanisms - if no section found, create empty section
            
            # Create empty section if none found
            if section is None:
                section = pd.Series()
            
            # Get mapped elements for processing
            mapped_elements = self._get_mapped_elements()
            
            # Process attachments and midspan data - pass connection type info
            result = self._process_attachments(node, section, mapped_elements, pole_scid, is_pole_to_reference)
            
            # Add basic connection information
            result['Pole'] = pole_scid
            result['To Pole'] = to_pole_scid
            
            # Get the initial span length and format it
            initial_span_distance = conn_info.get('span_distance', '')
            final_span_str = initial_span_distance

            # Special debugging for connections involving poles 10, 11, 12
            if any(pole in [pole_scid, to_pole_scid] for pole in ['10', '11', '12']):
                logging.info(f"  SPAN ASSIGNMENT DEBUG: {pole_scid} -> {to_pole_scid}: initial_span='{initial_span_distance}', connection_id='{conn_info.get('connection_id', '')}'")

            # If QC is active, apply tolerance to get the final span length
            if self.qc_reader and self.qc_reader.is_active():
                tolerance = self.config.get('processing_options', {}).get('span_length_tolerance', 3)
                qc_span_str = self.qc_reader.get_qc_span_length(pole_scid, to_pole_scid)
                final_span_str = self._apply_span_length_tolerance(initial_span_distance, qc_span_str, tolerance)
                logging.info(f"Span tolerance applied for {pole_scid} -> {to_pole_scid}: "
                             f"Excel='{initial_span_distance}', QC='{qc_span_str}', Final='{final_span_str}'")
                
                # Special debugging for connections involving poles 10, 11, 12
                if any(pole in [pole_scid, to_pole_scid] for pole in ['10', '11', '12']):
                    logging.info(f"  QC SPAN DEBUG: {pole_scid} -> {to_pole_scid}: QC_span='{qc_span_str}', final_span='{final_span_str}'")

            result['Span Length'] = self._format_span_distance(final_span_str)
            
            # Final span length assignment logging for debugging
            if any(pole in [pole_scid, to_pole_scid] for pole in ['10', '11', '12']):
                logging.info(f"  FINAL SPAN ASSIGNMENT: {pole_scid} -> {to_pole_scid}: result['Span Length']='{result['Span Length']}'")
            
            # Span length parsing (tension calculation removed)
            span_length_for_tension = self._parse_span_length(final_span_str)

            # Add pole information
            result['Address'] = self._get_pole_address(node)
            result['Pole Height/Class'] = self._format_pole_height_class(node)
            result['Existing Risers'] = self._count_existing_risers(node, pole_scid)
            
            # Add coordinates if available
            if node and 'latitude' in node:
                result['Latitude'] = Utils.round_coordinate(node['latitude'])
            if node and 'longitude' in node:
                result['Longitude'] = Utils.round_coordinate(node['longitude'])
            
            # Add pole tag from 'pole_tag_tagtext' column in nodes sheet
            if node and 'pole_tag_tagtext' in node:
                result['Pole Tag'] = node['pole_tag_tagtext']
            else:
                result['Pole Tag'] = ''  # Leave blank if not found
            
            # Add PDF report data (Structure Type, Existing Load, Proposed Load)
            if self.pdf_reader:
                try:
                    # Extract pole number from SCID (e.g., "001" -> 1)
                    pole_number = self._extract_pole_number_from_scid(pole_scid)
                    logging.debug(f"PoleDataProcessor: Extracting PDF data for SCID {pole_scid} -> pole number {pole_number}")
                    if pole_number:
                        pdf_data = self.pdf_reader.extract_pole_data(pole_number)
                        result['Structure Type'] = Utils.clean_structure_type(pdf_data.get('structure_type', ''))
                        result['Existing Load'] = pdf_data.get('existing_load', '')
                        result['Proposed Load'] = pdf_data.get('proposed_load', '')
                        logging.debug(f"PoleDataProcessor: Added PDF data for pole {pole_scid} (number {pole_number}): {pdf_data}")
                    else:
                        result['Structure Type'] = ''
                        result['Existing Load'] = ''
                        result['Proposed Load'] = ''
                        logging.debug(f"PoleDataProcessor: Could not extract pole number from SCID {pole_scid}")
                except Exception as e:
                    logging.error(f"PoleDataProcessor: Error extracting PDF data for pole {pole_scid}: {e}")
                    result['Structure Type'] = ''
                    result['Existing Load'] = ''
                    result['Proposed Load'] = ''
            else:
                logging.debug(f"PoleDataProcessor: No PDF reader available for pole {pole_scid}")
                result['Structure Type'] = ''
                result['Existing Load'] = ''
                result['Proposed Load'] = ''
            
            # Add guy information - use same priority order as Notes field
            guy_note = ''
            if node:
                # Use mr_note only
                mr_note = node.get('mr_note', '')
                guy_note = str(mr_note) if pd.notna(mr_note) else ''

            guy_info = self._extract_guy_info(guy_note)
            result['Guy Info'] = guy_info
            
            # Add Map field (can be customized based on requirements)
            result['Map'] = ''  # Empty for now, can be populated based on business logic
            
            # Add Line No. (will be set during output writing)
            result['Line No.'] = ''
            
            # Extract individual telecom provider heights from All_Comm_Heights
            all_comm_heights = result.get('All_Comm_Heights', '')
            
            # Initialize telecom provider fields - preserve existing values from attachment processing
            if 'Proposed MetroNet' not in result:
                result['Proposed MetroNet'] = ''
            if 'Verizon' not in result:
                result['Verizon'] = ''
            if 'AT&T' not in result:
                result['AT&T'] = ''
            if 'Comcast' not in result:
                result['Comcast'] = ''
            if 'Zayo' not in result:
                result['Zayo'] = ''
            if 'Jackson ISD' not in result:
                result['Jackson ISD'] = ''
            
            # Parse individual heights from All_Comm_Heights - only if not already populated
            # Skip provider-specific parsing for pole-to-reference connections
            if all_comm_heights and not is_pole_to_reference:
                import re
                # Pattern to match "height (provider)" format
                height_pattern = r"(\d+'\s*\d*\"?)\s*\(([^)]+)\)"
                matches = re.findall(height_pattern, all_comm_heights)
                
                for height, provider in matches:
                    # Clean up height formatting - ensure proper format
                    height = height.strip()
                    # Apply output formatting based on configuration
                    height = self._format_height_for_output(height)
                    
                    provider = provider.strip()
                    if 'MetroNet' in provider and not result['Proposed MetroNet']:
                        result['Proposed MetroNet'] = height
                    elif 'Verizon' in provider and not result['Verizon']:
                        result['Verizon'] = height
                    elif 'AT&T' in provider and not result['AT&T']:
                        result['AT&T'] = height
                    elif 'Comcast' in provider and not result['Comcast']:
                        result['Comcast'] = height
                    elif 'Zayo' in provider and not result['Zayo']:
                        result['Zayo'] = height
                    elif 'Jackson' in provider and not result['Jackson ISD']:
                        result['Jackson ISD'] = height
            elif all_comm_heights and is_pole_to_reference:
                # For pole-to-reference connections, skip provider-specific field population
                logging.debug(f"Skipping provider-specific field parsing for pole-to-reference connection (SCID {pole_scid})")
            
            # Tension calculation functionality has been removed
            result['Cable Type 1'] = ''
            result['Cable Diameter 1'] = ''
            result['Cable Type 2'] = ''
            result['Cable Diameter 2'] = ''
            result['Total Bundle Diameter'] = ''
            
            # Add guy fields from guy_info
            if guy_info and (guy_info['leads'] or guy_info['directions']):
                result['Guy Size'] = ', '.join(guy_info['sizes']) if guy_info['sizes'] else ''
                result['Guy Lead'] = ', '.join(guy_info['leads']) if guy_info['leads'] else ''
                result['Guy Direction'] = ', '.join(guy_info['directions']) if guy_info['directions'] else ''
            else:
                result['Guy Size'] = ''
                result['Guy Lead'] = ''
                result['Guy Direction'] = ''
            
            # Add notes field - check multiple possible note fields
            notes = ''
            if node:
                # Use mr_note only
                mr_note = node.get('mr_note', '')
                notes = str(mr_note) if pd.notna(mr_note) else ''

            result['Notes'] = notes
            
            # Add New Guy Required field based on MR Notes
            result['Guy Needed'] = self._determine_new_guy_required(notes)
            
            return result
            
        except Exception as e:
            logging.error(f"Error creating output row for {pole_scid} -> {to_pole_scid}: {e}")
            return None

    def _find_section(self, connection_id, sections_df, pole_scid=None, to_pole_scid=None):
        """Find section data for a connection_id, choosing section with lowest Proposed MetroNet height if multiple entries exist.
        If multiple rows match, further filter by pole_scid and to_pole_scid if columns exist."""
        if sections_df is None or sections_df.empty:
            return None
        
        # Filter for matching connection_id
        matching = sections_df[sections_df['connection_id'] == connection_id]
        
        # If possible, further filter by pole_scid and to_pole_scid
        pole_cols = [col for col in sections_df.columns if col.lower() in ['pole', 'from_pole', 'pole_scid', 'from_scid']]
        to_pole_cols = [col for col in sections_df.columns if col.lower() in ['to_pole', 'to_scid']]
        if pole_scid and to_pole_scid and not matching.empty:
            for pole_col in pole_cols:
                matching = matching[matching[pole_col] == pole_scid]
            for to_pole_col in to_pole_cols:
                matching = matching[matching[to_pole_col] == to_pole_scid]
        
        if matching.empty:
            return None
        
        if len(matching) == 1:
            return matching.iloc[0]
        
        # Choose entry with lowest overall attachment height when multiple entries exist
        if len(matching) > 1:
            matching_copy = matching.copy()
            
            # Find all height columns (POA_*HT)
            height_cols = [col for col in matching.columns if col.startswith("POA_") and col.endswith("HT")]
            
            if height_cols:
                # Calculate the lowest height for each row across all height columns
                min_heights = []
                for idx in matching.index:
                    row_heights = []
                    for ht_col in height_cols:
                        height_value = matching.loc[idx, ht_col]
                        if pd.notna(height_value):
                            try:
                                height_decimal = Utils.parse_height_decimal(height_value)
                                if height_decimal is not None:
                                    row_heights.append(height_decimal)
                            except:
                                continue
                    
                    # Find the minimum height for this row
                    min_height = min(row_heights) if row_heights else float('inf')
                    min_heights.append(min_height)
                
                # Find the row with the overall lowest height
                if min_heights and any(h != float('inf') for h in min_heights):
                    min_idx = min_heights.index(min(min_heights))
                    return matching.iloc[min_idx]
            
            # If no valid heights found, return first entry
            return matching.iloc[0]
        
        return matching.iloc[0]
    
    def _get_mapped_elements(self):
        """Get mapped providers and comm options from mapping data"""
        mapped = {
            'providers': set(),
            'midspan': set(),
            'comm_attach': set(),
            'comm_midspan': set()
        }
        
        for element, attribute, output in self.mapping_data:
            # Check if element is a telecom provider
            if element in self.config["telecom_providers"]:
                if attribute == "Attachment Ht":
                    mapped['providers'].add(element)
                elif attribute == "Midspan Ht":
                    mapped['midspan'].add(element)
            
            # Check if element is a comm field (comm1, comm2, comm3, comm4)
            if element in ["comm1", "comm2", "comm3", "comm4"]:
                if attribute == "Attachment Ht":
                    mapped['comm_attach'].add(element)
                elif attribute == "Midspan Ht":
                    mapped['comm_midspan'].add(element)
        
        return mapped
    
    def _process_attachments(self, node, section, mapped_elements, scid, is_pole_to_reference=False):
        """Process all attachment data for a pole"""
        # Initialize attachment dictionaries
        attach = {p: "" for p in mapped_elements['providers']}
        attach_midspan = {f"{p}_Midspan": "" for p in mapped_elements['midspan']}
        comm_attach = {c: "" for c in mapped_elements['comm_attach']}
        comm_midspan = {f"{c}_Midspan": "" for c in mapped_elements['comm_midspan']}
        
        # Track midspan heights per provider for grouping (similar to pole attachments)
        provider_midspan_heights = {}  # provider -> list of (height_decimal, height_formatted)
        
        # Add new fields for comprehensive communication data
        comm_attach['All_Comm_Heights'] = ""
        comm_attach['Total_Comm_Count'] = ""
        
        # Process attachment data from new file format
        telecom_heights = []
        power_heights = []
        power_midspan_heights = []
        all_telecom_attachments = []
        all_telecom_midspan = []
        
        # Get power and telecom attachments from attachment reader
        if self.attachment_reader:
            logging.debug(f"Processing attachments for pole {scid}")
            
            # Get power attachment (unchanged)
            power_attachment = self.attachment_reader.find_power_attachment(scid, self.config["power_keywords"])
            if power_attachment:
                # Store height, formatted height, and the keyword that matched
                keyword = power_attachment.get('keyword', '')
                power_heights.append((power_attachment['height_decimal'], power_attachment['height'], keyword))
                logging.info(f"Added power attachment for SCID {scid}: {power_attachment['height']} (keyword: {keyword})")
            else:
                logging.debug(f"No power attachment found for SCID {scid}")
            
            # Get power equipment (new functionality)
            power_equipment = self.attachment_reader.find_power_equipment(scid, self.config.get("power_equipment_keywords", []))
            if power_equipment:
                logging.info(f"Added power equipment for SCID {scid}: {power_equipment['equipment_count']} items")
            else:
                logging.debug(f"No power equipment found for SCID {scid}")
            
            # Process ALL communication attachments from raw SCID data directly (unified approach)
            raw_scid_data = self.attachment_reader.get_scid_data(scid)
            if not raw_scid_data.empty:
                logging.debug(f"Processing raw SCID data for {scid}: {len(raw_scid_data)} rows")
                logging.debug(f"Raw SCID data columns: {list(raw_scid_data.columns)}")
                logging.debug(f"Raw SCID data sample: {raw_scid_data[['company', 'measured', 'height_in_inches']].head().to_dict('records')}")
                # Look for ALL communication-related entries using configurable keywords
                comm_keywords = self.config.get("comm_keywords", ['catv com', 'telco com', 'fiber optic com', 'insulator', 'power guy', 'communication', 'comm'])
                logging.debug(f"Using communication keywords: {comm_keywords}")
                processed_attachments = {}  # Track by provider for provider-specific fields
                
                for _, row in raw_scid_data.iterrows():
                    measured = str(row.get('measured', '')).lower()
                    company = str(row.get('company', '')).lower()
                    
                    # Check if this is a communication attachment
                    # 1. Check if communication keywords match measured field (case-insensitive)
                    # Support wildcard (*) for substring matching or exact match
                    measured_clean = str(row.get('measured', '')).strip().lower()
                    is_comm_by_keyword = False
                    for kw in comm_keywords:
                        kw_clean = kw.strip().lower()
                        if kw_clean == 'guy':
                            # Special case: Guy always exact match
                            if kw_clean == measured_clean:
                                is_comm_by_keyword = True
                                break
                        elif kw_clean.endswith('*'):
                            # Wildcard match: substring
                            if kw_clean[:-1] in measured_clean:
                                is_comm_by_keyword = True
                                break
                        else:
                            # Exact match
                            if kw_clean == measured_clean:
                                is_comm_by_keyword = True
                                break
                    
                    # Debug: Show detailed matching for each keyword
                    if measured_clean:  # Only debug if there's actual measured data
                        logging.debug(f"Checking measured text: '{row.get('measured', '')}' -> cleaned: '{measured_clean}'")
                        for kw in comm_keywords:
                            kw_clean = kw.strip().lower()
                            # Check for wildcard or exact match
                            if kw_clean == 'guy':
                                match_result = kw_clean == measured_clean
                            elif kw_clean.endswith('*'):
                                match_result = kw_clean[:-1] in measured_clean
                            else:
                                match_result = kw_clean == measured_clean
                            logging.debug(f"  Keyword '{kw}' -> '{kw_clean}' == '{measured_clean}' ? {match_result}")
                    
                    if is_comm_by_keyword:
                        logging.debug(f"Communication keyword match: '{row.get('measured', '')}' -> '{measured_clean}' matches keywords: {comm_keywords}")
                    
                    # 2. Check if company name matches telecom providers or their synonyms
                    is_comm_by_company = self._is_telecom_company(company)
                    
                    # 3. Check if this is the proposed company (should be excluded from comm1, comm2, etc.)
                    proposed_company = self.config.get("proposed_company", "").strip().lower()
                    is_proposed_company = proposed_company and proposed_company in company
                    
                    # BOTH conditions must be met: comm keyword in measured AND telecom company
                    # BUT exclude proposed company from communication height processing
                    is_comm = is_comm_by_keyword and is_comm_by_company and not is_proposed_company
                    
                    if is_comm:
                        logging.debug(f"Found communication attachment: company='{row.get('company', '')}', measured='{row.get('measured', '')}', is_comm_by_keyword={is_comm_by_keyword}, is_comm_by_company={is_comm_by_company}, is_proposed_company={is_proposed_company}")
                    
                    if is_comm and 'height_in_inches' in row:
                        try:
                            height_value = row['height_in_inches']
                            if pd.isna(height_value) or str(height_value).strip() == '':
                                continue
                                
                            # Clean and convert height value
                            height_str = str(height_value).replace('"', '').replace('″', '').strip()
                            try:
                                height_inches = float(pd.to_numeric(height_str, errors='coerce'))
                                if not pd.isna(height_inches) and height_inches > 0:
                                    height_formatted = Utils.inches_to_feet_format(str(int(height_inches)))
                                    if height_formatted:  # Only proceed if conversion was successful
                                        # Apply output formatting based on configuration
                                        height_formatted = self._format_height_for_output(height_formatted)
                                        height_decimal = height_inches / 12
                                        
                                        # Determine provider for this attachment
                                        provider = None
                                        company_str = str(row.get('company', '')).strip()
                                        
                                        # Match to configured telecom providers or fall back to company name
                                        provider = self._match_telecom_provider(company_str)
                                        if not provider:
                                            provider = company_str if company_str else ""
                                        
                                        # Add to provider-specific processing (for provider fields)
                                        # Process attachments for both pole-to-pole AND pole-to-reference connections
                                        # Allow all communication companies to be processed, not just mapped providers
                                        if provider:  # Remove restriction to mapped_elements['providers']
                                            if provider not in processed_attachments:
                                                processed_attachments[provider] = []
                                            processed_attachments[provider].append((height_decimal, height_formatted))
                                        
                                        # Add to comprehensive list for All_Comm_Heights
                                        provider_info = f"{company_str} - {row.get('measured', '')}" if company_str else row.get('measured', '')
                                        entry = (height_decimal, height_formatted, provider_info)
                                        
                                        # Check if this exact height is already captured
                                        existing_heights = [x[0] for x in all_telecom_attachments]
                                        height_exists = any(abs(height_decimal - existing_height) < 0.01 for existing_height in existing_heights)
                                        
                                        if not height_exists:
                                            # Exclude Proposed MetroNet from comm1-4 columns
                                            if not self._match_metronet(provider):
                                                all_telecom_attachments.append(entry)
                                                telecom_heights.append(height_decimal)
                                                logging.debug(f"Added comm attachment for SCID {scid}: {height_formatted} ({provider_info})")
                                            else:
                                                logging.debug(f"Excluded Proposed MetroNet from comm columns for SCID {scid}: {height_formatted} ({provider_info})")
                                    else:
                                        logging.warning(f"Failed to format height {height_inches} for SCID {scid}")
                                else:
                                    logging.debug(f"Invalid height value {height_value} for SCID {scid}")
                            except Exception as e:
                                logging.warning(f"Error processing comm attachment for SCID {scid}: {e}")
                        except Exception as e:
                            logging.warning(f"Error processing comm attachment for SCID {scid}: {e}")
                
                # Process provider-specific attachments and assign to comm1, comm2, etc.
                # Process attachments for both pole-to-pole AND pole-to-reference connections
                
                # Collect all providers with their highest heights for sorting
                # Exclude Proposed MetroNet and its synonyms from comm1-4 fields
                provider_heights = []
                for provider, height_list in processed_attachments.items():
                    if height_list:
                        # Skip Proposed MetroNet and its synonyms - they should appear in their own column
                        if self._match_metronet(provider):
                            logging.debug(f"Excluded Proposed MetroNet from comm fields: {provider}")
                            continue
                        
                        # Sort heights from highest to lowest for this provider
                        height_list.sort(key=lambda x: x[0], reverse=True)
                        # Get the highest height for this provider
                        highest_height = height_list[0][0]  # height_decimal
                        # Combine all heights for this provider (deduplicate identical heights)
                        unique_heights = []
                        seen_heights = set()
                        for h in height_list:
                            if h[1] not in seen_heights:
                                unique_heights.append(h[1])
                                seen_heights.add(h[1])
                        combined_heights = ', '.join(unique_heights)
                        provider_heights.append((highest_height, provider, combined_heights))
                
                # Sort providers by their highest height (descending order)
                provider_heights.sort(key=lambda x: x[0], reverse=True)
                
                # Assign to comm fields based on height order
                comm_field_index = 1
                for highest_height, provider, combined_heights in provider_heights:
                    comm_field = f"comm{comm_field_index}"
                    attach[comm_field] = combined_heights
                    # Also populate comm_attach dictionary for the final output
                    if comm_field in comm_attach:
                        comm_attach[comm_field] = combined_heights
                    logging.info(f"Set {comm_field} ({provider}) attachment for SCID {scid}: {combined_heights} (highest: {highest_height:.2f}')")
                    
                    comm_field_index += 1
                    
                    # Stop if we've reached comm4 (assuming max 4 comm fields)
                    if comm_field_index > 4:
                        break
                
                # Also populate individual provider fields (including Proposed MetroNet)
                # First, collect all Proposed MetroNet synonyms into a single field
                metronet_heights = []
                logging.debug(f"Processing attachments for MetroNet collection. Available providers: {list(processed_attachments.keys())}")
                for provider, height_list in processed_attachments.items():
                    logging.debug(f"Checking provider '{provider}' with heights: {height_list}")
                    if height_list and self._match_metronet(provider):
                        metronet_heights.extend(height_list)
                        logging.debug(f"Collected Proposed MetroNet heights from {provider}: {height_list}")
                    else:
                        logging.debug(f"Provider '{provider}' not matched as MetroNet or has no heights")
                
                # Populate Proposed MetroNet field with all synonyms combined
                if metronet_heights and 'Proposed MetroNet' in attach:
                    metronet_heights.sort(key=lambda x: x[0], reverse=True)
                    # Deduplicate identical heights for MetroNet
                    unique_metronet_heights = []
                    seen_metronet_heights = set()
                    for h in metronet_heights:
                        if h[1] not in seen_metronet_heights:
                            unique_metronet_heights.append(h[1])
                            seen_metronet_heights.add(h[1])
                    combined_metronet_heights = ', '.join(unique_metronet_heights)
                    attach['Proposed MetroNet'] = combined_metronet_heights
                    logging.info(f"Set Proposed MetroNet attachment for SCID {scid}: {combined_metronet_heights}")
                
                # Populate other individual provider fields (excluding Proposed MetroNet synonyms)
                for provider, height_list in processed_attachments.items():
                    if height_list and provider in attach and not self._match_metronet(provider):
                        # Sort heights from highest to lowest for this provider
                        height_list.sort(key=lambda x: x[0], reverse=True)
                        # Combine all heights for this provider (deduplicate identical heights)
                        unique_heights = []
                        seen_heights = set()
                        for h in height_list:
                            if h[1] not in seen_heights:
                                unique_heights.append(h[1])
                                seen_heights.add(h[1])
                        combined_heights = ', '.join(unique_heights)
                        attach[provider] = combined_heights
                        logging.info(f"Set {provider} attachment for SCID {scid}: {combined_heights}")
        else:
            logging.warning("No attachment reader available - attachment data will not be processed")
        
        # Process section data for midspan - for both pole-to-pole and pole-to-reference connections
        midspan_processed_count = 0
        for col in section.index:
            if col.startswith("POA_") and not col.endswith("HT"):
                owner = str(section[col])
                ht_col = f"{col}HT"
                if ht_col in section and pd.notna(section[ht_col]):
                    fmt = Utils.parse_height_format(section[ht_col])
                    dec = Utils.parse_height_decimal(section[ht_col])
                    
                    self._process_midspan(owner, fmt, dec, attach_midspan, 
                                        power_midspan_heights, all_telecom_midspan, mapped_elements, provider_midspan_heights)
                    midspan_processed_count += 1
                    connection_type = "pole-to-reference" if is_pole_to_reference else "pole-to-pole"
                    logging.debug(f"Processed midspan data for {connection_type} connection (SCID {scid}), {col}: owner='{owner}', height={fmt}")
        
        if midspan_processed_count > 0:
            logging.debug(f"Total midspan entries processed for SCID {scid}: {midspan_processed_count}")
        else:
            logging.debug(f"No midspan data found in section for SCID {scid}")
        
        # Log connection type for debugging
        connection_type = "pole-to-reference" if is_pole_to_reference else "pole-to-pole"
        logging.debug(f"Processed {connection_type} connection (SCID {scid}): {midspan_processed_count} midspan entries")
        
        # Group provider midspan heights exactly like pole attachments (deduplicate and combine)
        # First, collect all Proposed MetroNet synonyms into a single field (like pole attachments)
        metronet_midspan_heights = []
        for provider, height_list in provider_midspan_heights.items():
            if height_list and self._match_metronet(provider):
                metronet_midspan_heights.extend(height_list)
        
        # Populate Proposed MetroNet midspan field with all synonyms combined
        if metronet_midspan_heights and "Proposed MetroNet" in mapped_elements['midspan']:
            metronet_midspan_heights.sort(key=lambda x: x[0], reverse=True)
            # Deduplicate identical heights for MetroNet midspan
            unique_metronet_heights = []
            seen_metronet_heights = set()
            for h in metronet_midspan_heights:
                if h[1] not in seen_metronet_heights:
                    unique_metronet_heights.append(h[1])
                    seen_metronet_heights.add(h[1])
            combined_metronet_midspan = ', '.join(unique_metronet_heights)
            attach_midspan["Proposed MetroNet_Midspan"] = combined_metronet_midspan
            logging.info(f"Set Proposed MetroNet midspan for SCID {scid}: {combined_metronet_midspan}")
        
        # Populate other individual provider midspan fields (excluding Proposed MetroNet synonyms)
        for provider, height_list in provider_midspan_heights.items():
            if height_list and provider in mapped_elements['midspan'] and not self._match_metronet(provider):
                # Sort heights from highest to lowest (like pole attachments)
                height_list.sort(key=lambda x: x[0], reverse=True)
                # Combine all heights for this provider (deduplicate identical heights)
                unique_heights = []
                seen_heights = set()
                for h in height_list:
                    if h[1] not in seen_heights:
                        unique_heights.append(h[1])
                        seen_heights.add(h[1])
                combined_heights = ', '.join(unique_heights)
                attach_midspan[f"{provider}_Midspan"] = combined_heights
                logging.info(f"Set {provider} midspan for SCID {scid}: {combined_heights}")
        
        # Group midspan heights by provider (exactly like pole attachments) before assigning to comm fields
        # This ensures multiple heights for the same provider are combined in the same comm field
        midspan_processed_by_provider = {}  # provider -> list of (height_decimal, height_formatted)
        
        # Group all_telecom_midspan by provider
        for height_decimal, height_formatted, owner in all_telecom_midspan:
            # Match provider from owner string
            matched_provider = self._match_telecom_provider(owner)
            if matched_provider:
                if matched_provider not in midspan_processed_by_provider:
                    midspan_processed_by_provider[matched_provider] = []
                midspan_processed_by_provider[matched_provider].append((height_decimal, height_formatted))
        
        # Collect all providers with their highest heights for sorting (like pole attachments)
        provider_midspan_heights_for_comm = []
        for provider, height_list in midspan_processed_by_provider.items():
            if height_list:
                # Sort heights from highest to lowest for this provider
                height_list.sort(key=lambda x: x[0], reverse=True)
                # Get the highest height for this provider
                highest_height = height_list[0][0]  # height_decimal
                # Combine all heights for this provider (deduplicate identical heights)
                unique_heights = []
                seen_heights = set()
                for h in height_list:
                    if h[1] not in seen_heights:
                        unique_heights.append(h[1])
                        seen_heights.add(h[1])
                combined_heights = ', '.join(unique_heights)
                provider_midspan_heights_for_comm.append((highest_height, provider, combined_heights))
        
        # Sort providers by their highest height (descending order)
        provider_midspan_heights_for_comm.sort(key=lambda x: x[0], reverse=True)
        
        # Assign to comm midspan fields based on height order (exactly like pole attachments)
        comm_field_index = 1
        for highest_height, provider, combined_heights in provider_midspan_heights_for_comm:
            comm_field = f"comm{comm_field_index}_Midspan"
            if comm_field in comm_midspan:
                comm_midspan[comm_field] = combined_heights
                logging.info(f"Set {comm_field} ({provider}) midspan for SCID {scid}: {combined_heights} (highest: {highest_height:.2f}')")
            
            comm_field_index += 1
            
            # Stop if we've reached comm4 (assuming max 4 comm fields)
            if comm_field_index > 4:
                break
        
        # Calculate power heights
        power_data = self._calculate_power_heights(power_heights, power_midspan_heights, telecom_heights)
        
        # Log final results for debugging
        if any(attach.values()) or any(power_data.values()) or comm_attach.get('All_Comm_Heights') or any(comm_midspan.values()):
            logging.info(f"Final attachments for SCID {scid}:")
            for key, value in {**attach, **power_data, **comm_attach, **comm_midspan}.items():
                if value:
                    logging.info(f"  {key}: {value}")
        
        # Add streetlight (bottom of bracket) height
        streetlight_from_find = self.attachment_reader.find_streetlight_attachment(scid) if self.attachment_reader else None
        
        # New: Find street light for power company, measured contains configured keywords
        street_light_height_processed = ""
        if self.attachment_reader:
            try:
                df_scid_data = self.attachment_reader.get_scid_data(scid)
                if not df_scid_data.empty:
                    street_keywords = self._get_street_light_keywords()
                    require_power_company = self._keywords_require_power_company(street_keywords)
                    power_company_config = self.config.get("power_company", "").strip().lower()

                    if not power_company_config and require_power_company:
                        logging.debug(f"SCID {scid}: Power company not configured. Skipping street light detection for riser keywords.")
                    else:
                        df_filtered = df_scid_data.copy()
                        if power_company_config:
                            df_filtered['company_stripped'] = df_filtered['company'].astype(str).str.strip().str.lower()
                            power_company_pattern = r'\b' + re.escape(power_company_config) + r'\b'
                            company_mask = df_filtered['company_stripped'].str.contains(power_company_pattern, na=False, regex=True)
                            if require_power_company:
                                df_filtered = df_filtered[company_mask]
                            else:
                                blank_mask = df_filtered['company_stripped'].eq('')
                                df_filtered = df_filtered[company_mask | blank_mask]

                        if not df_filtered.empty:
                            df_filtered = df_filtered.copy()
                            df_filtered['measured_stripped'] = df_filtered['measured'].astype(str).str.strip().str.lower()

                            keyword_pattern = self._build_keyword_regex(street_keywords)
                            if keyword_pattern:
                                street_rows = df_filtered[df_filtered['measured_stripped'].str.contains(keyword_pattern, na=False, regex=True)]
                            else:
                                street_rows = df_filtered[df_filtered['measured_stripped'].str.contains('street', na=False)]

                            if not street_rows.empty:
                                street_rows = street_rows.copy()
                                street_rows['company_stripped'] = street_rows['company'].astype(str).str.strip().str.lower()
                                if power_company_config:
                                    company_mask = street_rows['company_stripped'].str.contains(
                                        power_company_pattern, na=False, regex=True
                                    )
                                else:
                                    company_mask = pd.Series(False, index=street_rows.index)

                                requires_mask = street_rows['measured_stripped'].apply(
                                    lambda text: self._measurement_requires_power_company(text, street_keywords)
                                )
                                street_rows = street_rows[
                                    ~(requires_mask & ~company_mask)
                                ]

                                if not street_rows.empty:
                                    s_rows_copy = street_rows.copy()
                                    s_rows_copy['height_numeric'] = pd.to_numeric(
                                        s_rows_copy['height_in_inches'].astype(str).str.replace('"', '').str.replace('″', ''),
                                        errors='coerce'
                                    )
                                    s_rows_copy = s_rows_copy.dropna(subset=['height_numeric'])
                                    if not s_rows_copy.empty:
                                        min_row = s_rows_copy.loc[s_rows_copy['height_numeric'].idxmin()]
                                        street_light_height_processed = Utils.inches_to_feet_format(str(int(min_row['height_numeric'])))
                                        street_light_height_processed = self._format_height_for_output(street_light_height_processed)
            except Exception as e:
                logging.error(f"Error processing street light height for SCID {scid}: {e}")
        
        result = {**attach, **attach_midspan, **comm_attach, **comm_midspan, **power_data}
        
        # Add power equipment data
        if self.attachment_reader:
            power_equipment = self.attachment_reader.find_power_equipment(scid, self.config.get("power_equipment_keywords", []))
            if power_equipment and power_equipment.get('equipment_list'):
                result['Power Equipments'] = power_equipment['equipment_list']
            else:
                result['Power Equipments'] = ''
                logging.debug(f"No Power Equipment data for SCID {scid}")
        else:
            result['Power Equipments'] = ''
            logging.debug(f"No attachment reader available for Power Equipment extraction")
        
        # This is for the pre-existing field 'Streetlight (bottom of bracket)'
        if streetlight_from_find:
            result['Streetlight (bottom of bracket)'] = streetlight_from_find['height']
        else:
            result['Streetlight (bottom of bracket)'] = ''
        
        # This is for the new field 'Street Light Height'
        result['Street Light Height'] = street_light_height_processed if street_light_height_processed else ''
        
        return result
    
    def _process_midspan(self, owner, fmt, dec, attach_midspan, 
                        power_midspan_heights, all_telecom_midspan, mapped_elements, provider_midspan_heights=None):
        """Process midspan data"""
        # Apply output formatting based on configuration
        formatted_fmt = self._format_height_for_output(fmt) if fmt else fmt
        
        # MetroNet midspan - collect heights instead of overwriting
        if self._match_metronet(owner):
            if "Proposed MetroNet" in mapped_elements['midspan']:
                if provider_midspan_heights is not None:
                    if "Proposed MetroNet" not in provider_midspan_heights:
                        provider_midspan_heights["Proposed MetroNet"] = []
                    if dec is not None and formatted_fmt:
                        provider_midspan_heights["Proposed MetroNet"].append((dec, formatted_fmt))
        
        # Telecom midspan - collect heights instead of overwriting
        matched = self._match_telecom_provider(owner)
        if matched and matched in mapped_elements['midspan']:
            if provider_midspan_heights is not None:
                if matched not in provider_midspan_heights:
                    provider_midspan_heights[matched] = []
                if dec is not None and formatted_fmt:
                    provider_midspan_heights[matched].append((dec, formatted_fmt))
        
        # Collect for comm sorting (exclude Proposed MetroNet from comm1-4 columns)
        if matched or self._match_metronet(owner):
            if dec is not None and fmt:
                # Only add to comm columns if it's not Proposed MetroNet
                if not self._match_metronet(owner):
                    all_telecom_midspan.append((dec, formatted_fmt, owner))
                else:
                    logging.debug(f"Excluded Proposed MetroNet midspan from comm columns: {formatted_fmt} ({owner})")
        
        # Power midspan
        if any(kw.lower() in owner.lower() for kw in self.config["power_keywords"]):
            if dec is not None:
                power_midspan_heights.append((dec, formatted_fmt))

    def _assign_comm_attachments(self, telecom_data, comm_dict, mapped_comms):
        """Sort and assign telecom attachments to comm1-4, and capture ALL communication heights"""
        logging.debug(f"Raw telecom data before filtering: {telecom_data}")
        logging.debug(f"Mapped comms: {mapped_comms}")
        logging.debug(f"Comm dict keys: {list(comm_dict.keys())}")
        
        # Filter out entries without measured data
        filtered_telecom_data = [x for x in telecom_data if x[2]]
        
        # Determine if this is attachment data or midspan data based on comm_dict keys
        is_midspan_data = any(key.endswith('_Midspan') for key in comm_dict.keys())
        
        # Updated filter to include expanded communication keywords
        # Include: 'CATV Com', 'Telco Com', 'Fiber Optic Com', 'insulator', 'Power Guy'
        # NOTE: Keyword filtering applies only to attachment data, not midspan data
        if not is_midspan_data:
            keywords = self.config.get("comm_keywords", ['catv com', 'telco com', 'fiber optic com', 'insulator', 'power guy', 'catv', 'telco', 'fiber', 'communication', 'comm'])
            
            # Extract measured column data for exact matching (case-insensitive)
            def extract_measured_data(provider_info):
                """Extract measured data from provider_info string"""
                if ' - ' in provider_info:
                    # Format: "Company - measured_data"
                    return provider_info.split(' - ', 1)[1].strip()
                else:
                    # Format: "measured_data" (no company)
                    return provider_info.strip()
            
            # Use exact text matching instead of substring matching
            keyword_filtered = []
            for x in filtered_telecom_data:
                measured_data = extract_measured_data(str(x[2]))
                if any(kw.lower() == measured_data.lower() for kw in keywords):
                    keyword_filtered.append(x)
            
            # If keyword matches are present, use them; otherwise, use the full filtered list
            # This ensures we don't accidentally exclude valid communication attachments
            if keyword_filtered:
                filtered_telecom_data = keyword_filtered
                logging.debug(f"Using keyword-filtered data: {len(keyword_filtered)} entries")
            else:
                logging.debug(f"No keyword matches found, using all filtered data: {len(filtered_telecom_data)} entries")
                # Log what we're including when no keyword matches are found
                if filtered_telecom_data:
                    logging.debug(f"Non-keyword filtered entries: {[x[2] for x in filtered_telecom_data[:5]]}")  # Log first 5
        else:
            logging.debug(f"Skipping keyword filtering for midspan data: {len(filtered_telecom_data)} entries")
        
        # Sort data (highest to lowest)
        filtered_telecom_data.sort(key=lambda x: x[0], reverse=True)
        logging.debug(f"Sorted telecom data for comm assignment: {filtered_telecom_data}")
        
        # Assign to comm fields (first 4 only)
        comm_names = ["comm1", "comm2", "comm3", "comm4"]
        for i, comm in enumerate(comm_names):
            if is_midspan_data:
                # For midspan data, always use the _Midspan suffix
                key = f"{comm}_Midspan"
            else:
                # For attachment data, use the base comm name (comm1, comm2, etc.)
                key = comm
            
            logging.debug(f"Checking {comm}: key={key}, key in comm_dict={key in comm_dict}, is_midspan_data={is_midspan_data}")
            if key in comm_dict and i < len(filtered_telecom_data):
                comm_dict[key] = filtered_telecom_data[i][1]
                logging.info(f"Assigned {filtered_telecom_data[i][1]} to {key}")
            else:
                logging.debug(f"Skipping assignment for {comm}: key={key}, key in comm_dict={key in comm_dict}, i={i}, data_len={len(filtered_telecom_data)}")
        
        # NEW: Create comprehensive summary of ALL communication attachment heights (only for attachment data, not midspan)
        if filtered_telecom_data and not is_midspan_data:
            all_comm_heights = []
            for height_decimal, height_formatted, provider in filtered_telecom_data:
                # Include provider info if available
                if provider and str(provider).strip():
                    all_comm_heights.append(f"{height_formatted} ({provider})")
                else:
                    all_comm_heights.append(height_formatted)
            
            # Add comprehensive field to comm_dict
            comm_dict['All_Comm_Heights'] = '; '.join(all_comm_heights)
            logging.info(f"All communication heights captured: {comm_dict['All_Comm_Heights']}")
            
            # Also add count of total communication attachments
            comm_dict['Total_Comm_Count'] = str(len(filtered_telecom_data))
            logging.info(f"Total communication attachments found: {len(filtered_telecom_data)}")
        else:
            logging.debug("No filtered telecom data found for assignment or this is midspan data")
    
    def _calculate_power_heights(self, power_heights, power_midspan_heights, telecom_heights):
        """Calculate lowest power heights"""
        lowest_power = ""
        lowest_power_midspan = ""
        lowest_power_type = ""
        
        if power_heights:
            # Filter out invalid heights (0.00 and negative heights are invalid for power attachments)
            # power_heights contains (height_decimal, height_formatted, keyword) tuples
            # Handle both 2-element (legacy) and 3-element (with keyword) tuples
            valid_heights = []
            for entry in power_heights:
                if len(entry) >= 2 and entry[0] > 0:
                    # Extract height_decimal, height_formatted, and keyword (if available)
                    h = entry[0]
                    f = entry[1]
                    k = entry[2] if len(entry) > 2 else ""
                    valid_heights.append((h, f, k))
            
            if valid_heights:
                min_threshold = max(telecom_heights) if telecom_heights else 0
                valid_power = [(h, f, k) for h, f, k in valid_heights if h >= min_threshold]
                
                if valid_power:
                    lowest_entry = min(valid_power, key=lambda x: x[0])
                    lowest_power = lowest_entry[1]
                    lowest_power_type = lowest_entry[2] if lowest_entry[2] else ""
                else:
                    lowest_entry = min(valid_heights, key=lambda x: x[0])
                    lowest_power = lowest_entry[1]
                    lowest_power_type = lowest_entry[2] if lowest_entry[2] else ""
        
        if power_midspan_heights:
            lowest_power_midspan = min(power_midspan_heights, key=lambda x: x[0])[1]
        
        return {
            'Power Height': lowest_power,
            'Power Midspan': lowest_power_midspan,
            'Power Type': lowest_power_type
        }
    
    def _match_metronet(self, owner):
        """Check if owner matches Proposed MetroNet (case insensitive) 
           For 'power guy' keyword, still requires company/owner name to be present."""
        owner_str = str(owner).lower()
        keywords = self._get_proposed_company_keywords()
        
        logging.debug(f"_match_metronet called with owner='{owner}', owner_str='{owner_str}', keywords={keywords}")
        
        # Check regular MetroNet keywords first
        for keyword in keywords:
            if keyword.lower() in owner_str:
                logging.debug(f"Matched MetroNet keyword '{keyword}' in '{owner_str}'")
                return True
        
        # Special handling for 'power guy' - must have company name present
        if "power guy" in owner_str:
            # Check if any company/provider names are also present
            all_providers = set()
            telecom_keywords = self.config.get("telecom_keywords", {})
            if telecom_keywords:
                for provider_keywords in telecom_keywords.values():
                    all_providers.update([k.lower() for k in provider_keywords])
            else:
                all_providers.update([p.lower() for p in self.config.get("telecom_providers", []) if p])
                all_providers.update(self._get_proposed_company_keywords())
            
            # Also check power company names
            power_keywords = [k.lower() for k in self.config.get("power_keywords", [])]
            all_providers.update(power_keywords)
            
            # If any provider/company name is found along with 'power guy', it's valid
            for provider in all_providers:
                if provider in owner_str and provider != "power guy":
                    return True
        
        return False
    
    def _match_telecom_provider(self, owner):
        """Match owner to telecom provider (case insensitive)"""
        owner_str = str(owner).lower()
        telecom_keywords = self.config.get("telecom_keywords", {})
        for provider, keywords in telecom_keywords.items():
            if any(k.lower() in owner_str for k in keywords):
                return provider
        
        # Fall back to proposed company name if telecom keywords are absent
        if any(keyword in owner_str for keyword in self._get_proposed_company_keywords()):
            return "Proposed MetroNet"
        
        # As a last resort, try matching provider names directly from telecom_providers
        for provider in self.config.get("telecom_providers", []):
            if provider and provider.lower() in owner_str:
                return provider
        
        return None

    def _get_proposed_company_keywords(self):
        """Return a set of normalized keywords for the proposed company (MetroNet)."""
        keywords = set()
        proposed_company = self.config.get("proposed_company", "").strip().lower()
        
        if proposed_company:
            keywords.add(proposed_company)
            
            # Include version without leading 'proposed ' if present
            if proposed_company.startswith("proposed "):
                keywords.add(proposed_company[len("proposed "):])
            
            # Include alphanumeric-only variant
            normalized = re.sub(r'[^a-z0-9]+', ' ', proposed_company).strip()
            if normalized:
                keywords.add(normalized)
        
        # Ensure legacy MetroNet keywords still match even if proposed company differs
        keywords.update({"proposed metronet", "metronet", "proposed mnt", "mnt"})
        
        return [kw for kw in keywords if kw]

    def _get_street_light_keywords(self):
        """Get configured keywords for identifying street light measurements."""
        configured = self.config.get("street_light_keywords", [])
        keywords = [kw.strip().lower() for kw in configured if isinstance(kw, str) and kw.strip()]
        if not keywords:
            keywords = ["street"]
        return keywords

    @staticmethod
    def _measurement_requires_power_company(measured_text, keywords):
        """Check if measured text contains a keyword that requires power company validation."""
        measured_lower = str(measured_text).strip().lower()
        for keyword in keywords:
            if not keyword:
                continue
            keyword_lower = keyword.strip().lower()
            if 'riser' in keyword_lower and keyword_lower in measured_lower:
                return True
        return False

    @staticmethod
    def _keywords_require_power_company(keywords):
        """Determine if keyword list requires power company context (e.g., contains 'riser')."""
        for kw in keywords:
            if isinstance(kw, str) and 'riser' in kw.lower():
                return True
        return False

    @staticmethod
    def _build_keyword_regex(keywords):
        """Build a regex pattern supporting '*' wildcard for given keywords."""
        patterns = []
        for kw in keywords:
            escaped = re.escape(kw).replace(r'\*', '.*')
            patterns.append(escaped)
        if not patterns:
            return None
        return r'(?:' + '|'.join(patterns) + r')'
    
    def _get_pole_address(self, node):
        """Get pole address from geocoding cache or service"""
        lat, lon = node.get('latitude'), node.get('longitude')
        if not lat or not lon:
            return ''
            
        if self.geocoder:
            # If geocoding is enabled, try cache first, then geocoding service
            address = self.geocoder.reverse(lat, lon)
            return address
        else:
            # If geocoding is disabled, only check cache
            try:
                key = f"{round(float(lat), 7)},{round(float(lon), 7)}"
                return self.geocoder.cache.get(key, '')
            except:
                return ''
    
    def _format_span_distance(self, span_distance):
        """Format span distance as rounded feet with ' suffix"""
        if not span_distance:
            return ""
        
        try:
            # Convert to float and round to nearest whole number
            distance_feet = round(float(span_distance))
            return f"{distance_feet}'"
        except (ValueError, TypeError):
            # If conversion fails, return original value
            return str(span_distance)

    def _format_pole_height_class(self, node):
        """Format pole height and class from pole_spec column"""
        pole_spec = node.get('pole_spec', '')
        
        if pole_spec:
            # Parse pole_spec format like '35-4 SOUTHERN PINE (NESC Standard)' to extract '35/4'
            try:
                # Extract height-class from the beginning of the string
                # Pattern: digits-digits followed by space or end
                import re
                match = re.match(r'^(\d+)-(\d+)', str(pole_spec).strip())
                if match:
                    height = match.group(1)
                    pole_class = match.group(2)
                    return f"{height}/{pole_class}"
                else:
                    # If no match found, return blank
                    return ""
            except Exception as e:
                logging.debug(f"Error parsing pole_spec '{pole_spec}': {e}")
                return ""
        return ""

    def _count_existing_risers(self, node, scid=None):
        """Count existing risers from attachment data, excluding MetroNet"""
        # Use attachment data if available
        if self.attachment_reader and scid:
            try:
                count = self.attachment_reader.count_existing_risers_from_attachments(scid)
                logging.debug(f"Counted {count} existing risers from attachment data for SCID {scid}")
                return str(count)
            except Exception as e:
                logging.error(f"Error counting risers from attachment data for SCID {scid}: {e}")
                # Fall back to node data if attachment data fails
                pass
        
        # Fallback to node data (original logic)
        count = 0
        for key, val in node.items():
            if key.startswith("POA_") and not key.endswith("HT"):
                if isinstance(val, str):
                    owner = val.lower()
                    if "riser" in owner and not self._match_metronet(val):
                        count += 1
        return str(count)
    
    
    def _parse_span_length(self, span_distance):
        """Parse span length from span distance string"""
        try:
            if not span_distance:
                return None
                
            # Convert to string and clean
            span_str = str(span_distance).strip()
            if not span_str or span_str.lower() in ['nan', 'none', '']:
                return None
            
            # Remove common suffixes like ' or ft
            span_str = span_str.replace("'", "").replace("ft", "").replace("feet", "").strip()
            
            # Try to convert to float
            return float(span_str)
            
        except (ValueError, TypeError):
            logging.warning(f"Could not parse span length: {span_distance}")
            return None
    
    def _parse_height_value(self, value):
        """Parse height value from various formats including feet'inches" """
        try:
            # Handle NaN/None values early
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return None

            # Use shared utility to parse into decimal-feet
            decimal_feet = Utils.parse_height_decimal(value)

            if decimal_feet is not None:
                # Round to two decimals for consistency
                return round(decimal_feet, 2)

            logging.debug(f"Utils.parse_height_decimal was unable to parse '{value}' – no fallback mechanism")

            # Single data source: Only use properly formatted height data
            # Return None for unparseable values
            return None

        except (ValueError, TypeError) as e:
            logging.warning(f"Could not parse height value '{value}': {e}")
            return None
    
    def _format_height_for_output(self, height_str):
        """Format height string based on output_decimal configuration"""
        try:
            if not height_str or str(height_str).strip() == '':
                return ''
            
            # Check if output_decimal is enabled
            output_decimal = self.config.get("processing_options", {}).get("output_decimal", False)
            
            if output_decimal:
                # Convert to decimal format
                return Utils.feet_inches_to_decimal_format(height_str)
            else:
                # Return original format
                return str(height_str)
                
        except Exception as e:
            logging.warning(f"Error formatting height '{height_str}': {e}")
            return str(height_str)
    
    def _convert_measurement_for_alden(self, value):
        """
        Convert decimal measurements to Alden format for Alden sheet.
        Examples: 22.08 -> '22ft 1in', 23.58 -> '23ft 7in'
        Handles multiple comma-separated heights: "22.08, 23.58" -> "22ft 1in, 23ft 7in"
        
        Args:
            value: The value to convert (could be string, int, float, etc., or comma-separated values)
            
        Returns:
            Converted value in Alden format if it's a decimal measurement, None otherwise
            For multiple values, converts each and joins with comma
        """
        try:
            if value is None:
                return None
                
            value_str = str(value).strip()
            if not value_str or value_str == '':
                return None
            
            # Check if value contains multiple heights (comma-separated)
            if ',' in value_str:
                # Split by comma and convert each value
                parts = [part.strip() for part in value_str.split(',')]
                converted_parts = []
                original_format_used = False
                
                for part in parts:
                    if not part:
                        continue
                    
                    # Try to convert this part
                    part_converted = self._convert_single_measurement_for_alden(part)
                    if part_converted:
                        converted_parts.append(part_converted)
                    else:
                        # Keep original format if conversion fails
                        converted_parts.append(part)
                        original_format_used = True
                
                if converted_parts:
                    # If any conversions succeeded, return combined result
                    result = ', '.join(converted_parts)
                    if original_format_used:
                        logging.debug(f"Partial conversion for '{value_str}': some values kept original format")
                    return result
                else:
                    return None
            
            # Single value conversion
            return self._convert_single_measurement_for_alden(value_str)
            
        except Exception as e:
            logging.warning(f"Error converting measurement for Alden: {e}")
            return None
    
    def _convert_single_measurement_for_alden(self, value_str):
        """Convert a single measurement value to Alden format"""
        try:
            # Try to parse as a decimal number
            try:
                decimal_value = float(value_str)
                
                # Check if it's a reasonable measurement value (typically between 0 and 1000 feet)
                if 0 <= decimal_value < 1000:
                    # Use the utility to convert to Alden format
                    converted = Utils.decimal_feet_to_alden_format(decimal_value)
                    if converted:
                        logging.debug(f"Converted '{value_str}' to Alden format: '{converted}'")
                        return converted
                    
            except (ValueError, TypeError):
                # Not a decimal number, return None to keep original value
                pass
            
            return None
            
        except Exception as e:
            logging.warning(f"Error converting single measurement for Alden: {e}")
            return None
    
    def write_output(self, result_data, output_file):
        """Write processed data to Excel output file"""
        try:
            if not result_data:
                logging.warning("No data to write to output file")
                return

            # Filter out None or non-dict items before sorting
            filtered_data = [item for item in result_data if item and isinstance(item, dict)]
            if not filtered_data:
                logging.warning("No valid data to write after filtering")
                return

            # Sort data only if QC file is not active (preserve QC order when active)
            if self.qc_reader and self.qc_reader.is_active():
                # QC file is active - preserve the exact order from QC file
                sorted_data = filtered_data
                logging.info("QC file is active - preserving QC order for main sheet")
            else:
                # No QC file - sort by pole SCID as usual
                sorted_data = sorted(filtered_data, key=lambda x: Utils.extract_numeric_part(x.get('Pole', '')))
                logging.info("No QC file - sorting main sheet by pole SCID")

            # Create data cache for QC sheet population
            self._processed_data_cache = {}
            for row in sorted_data:
                pole = row.get('Pole', '').strip()
                if pole:
                    self._processed_data_cache[pole] = row

            # Check if output file exists, if not, try to create it from template
            from pathlib import Path
            output_path = Path(output_file)
            
            if not output_path.exists():
                # Try to find and copy template
                template_path = None
                if hasattr(self, 'config') and self.config:
                    # Look for template in common locations
                    possible_templates = [
                        'C:/Users/nsaro/Desktop/Test/Consumer SS Template.xltm',
                        'Consumer SS Template.xltm',
                        'template.xlsx',
                        'template.xltm'
                    ]
                    
                    for template in possible_templates:
                        if Path(template).exists():
                            template_path = template
                            break
                
                if template_path:
                    logging.info(f"Output file doesn't exist, copying from template: {template_path}")
                    import shutil
                    try:
                        shutil.copy2(template_path, output_file)
                        logging.info(f"Successfully created output file from template")
                    except Exception as e:
                        logging.error(f"Failed to copy template: {e}")
                        return
                else:
                    logging.error(f"Output file '{output_file}' doesn't exist and no template found")
                    return

            # Validate the output file after creation/copying
            if not output_path.exists() or output_path.stat().st_size == 0:
                logging.error(f"Output file '{output_file}' is missing or empty.")
                return

            # Attempt to load the workbook inside a try/except block to catch EOFError
            try:
                # Use keep_vba=True only for .xlsm files, not for .xlsx files
                output_path = Path(output_file)
                if output_path.suffix.lower() == '.xlsm':
                    wb = load_workbook(output_file, keep_vba=True)
                else:
                    wb = load_workbook(output_file)
            except EOFError as eof_error:
                logging.error(f"EOFError encountered when loading workbook '{output_file}': {eof_error}. The template file may be corrupted.")
                return
            except Exception as e:
                logging.error(f"Error loading workbook '{output_file}': {e}")
                return

            # Apply final span length tolerance check to all sheet data
            if self.qc_reader and self.qc_reader.is_active():
                logging.info("Applying final span length tolerance check to sheet data")
                tolerance = self.config.get('processing_options', {}).get('span_length_tolerance', 3)
                logging.info(f"Span length tolerance setting: {tolerance}")
                
                tolerance_updates = 0
                for row_data in sorted_data:
                    pole = row_data.get('Pole', '')
                    to_pole = row_data.get('To Pole', '')
                    
                    if pole and to_pole and not self._is_end_marker(to_pole):
                        # Get QC span length
                        qc_span = self.qc_reader.get_qc_span_length(pole, to_pole)
                        excel_span = row_data.get('Span Length', '')
                        
                        # Log every connection for debugging
                        if qc_span or excel_span:
                            logging.info(f"Tolerance check: {pole} -> {to_pole}: Excel='{excel_span}', QC='{qc_span}'")
                        
                        if qc_span and excel_span:
                            # Apply tolerance check
                            final_span = self._apply_span_length_tolerance(excel_span, qc_span, tolerance)
                            
                            # Always update with the final span length (tolerance method handles the logic)
                            row_data['Span Length'] = final_span
                            
                            # Log the result for debugging
                            if final_span != excel_span:
                                logging.info(f"Final tolerance update: {pole} -> {to_pole}: '{excel_span}' -> '{final_span}' (QC: {qc_span}, tolerance: {tolerance})")
                                tolerance_updates += 1
                            else:
                                logging.info(f"Excel span length retained: {pole} -> {to_pole}: '{excel_span}' (QC: {qc_span}, tolerance: {tolerance})")
                    
                    # Ensure END markers remain intact after any tolerance logic
                    self._apply_end_marker(row_data)
                
                logging.info(f"Completed span length tolerance check: {tolerance_updates} updates applied")
            
            # Write data to all sheets that have Pole/To Pole columns
            sheets_written = 0
            if hasattr(self, 'template_scids_by_sheet') and self.template_scids_by_sheet:
                # Write to each sheet found in template
                for sheet_name in self.template_scids_by_sheet.keys():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        
                        # Filter data to only include rows that match this sheet's Pole/To Pole pairs
                        sheet_scids = self.template_scids_by_sheet[sheet_name]
                        sheet_pole_to_pole_pairs = {(str(pole).strip(), str(to_pole).strip()) for pole, to_pole, _ in sheet_scids}
                        
                        # Filter sorted_data to only include rows matching this sheet
                        filtered_sorted_data = []
                        for row_data in sorted_data:
                            pole_val = str(row_data.get('Pole', '')).strip()
                            to_pole_val = str(row_data.get('To Pole', '')).strip()
                            if (pole_val, to_pole_val) in sheet_pole_to_pole_pairs:
                                filtered_sorted_data.append(row_data)
                        
                        logging.info(f"Filtered to {len(filtered_sorted_data)} rows for sheet '{sheet_name}' (from {len(sorted_data)} total rows)")
                        
                        # Write data; using mapped writing if available, else a simple write
                        if hasattr(self, 'mapping_data') and self.mapping_data:
                            self._write_data_to_worksheet(ws, filtered_sorted_data, self.mapping_data, sheet_name)
                        else:
                            self._write_data_simple(ws, filtered_sorted_data, sheet_name)
                        
                        sheets_written += 1
                    else:
                        logging.warning(f"Sheet '{sheet_name}' not found in output workbook")
            else:
                # Fallback to single sheet mode (backward compatibility)
                if hasattr(self, 'config') and self.config:
                    worksheet_name = self.config.get('output_settings', {}).get('worksheet_name', 'Consumers pg1')
                else:
                    worksheet_name = 'Consumers pg1'

                if worksheet_name in wb.sheetnames:
                    ws = wb[worksheet_name]
                else:
                    ws = wb.active
                    logging.warning(f"Worksheet '{worksheet_name}' not found, using '{ws.title}'")

                # Write data; using mapped writing if available, else a simple write
                if hasattr(self, 'mapping_data') and self.mapping_data:
                    self._write_data_to_worksheet(ws, sorted_data, self.mapping_data, worksheet_name)
                else:
                    self._write_data_simple(ws, sorted_data, worksheet_name)
                
                sheets_written = 1
            
            logging.info(f"Wrote data to {sheets_written} sheet(s)")

            # Automatically populate QC sheet if QC reader is active
            if self.qc_reader and self.qc_reader.is_active():
                logging.info("QC reader is active - populating QC sheet")
                self._populate_qc_sheet(wb)
                
                # Add conditional formatting to compare main sheet and QC sheet
                logging.info("Adding conditional formatting to compare main sheet and QC sheet")
                # Use first non-QC sheet for comparison
                main_sheet_name = None
                if hasattr(self, 'template_scids_by_sheet') and self.template_scids_by_sheet:
                    for sheet in self.template_scids_by_sheet.keys():
                        if sheet in wb.sheetnames and sheet != "QC":
                            main_sheet_name = sheet
                            break
                
                if main_sheet_name is None:
                    main_sheet_name = self.config.get('output_settings', {}).get('worksheet_name', 'Consumers pg1')
                
                self._add_sheet_comparison_formatting(wb, main_sheet_name)

            # Apply Alden QC comparison if Alden QC reader is active
            if self.alden_qc_reader and self.alden_qc_reader.is_active():
                logging.info("Alden QC reader is active - performing Alden comparison")
                self._apply_alden_qc_comparison(wb)
            else:
                logging.info("Alden QC reader not active - skipping Alden comparison")

            wb.save(output_file)
            logging.info(f"Successfully wrote {len(sorted_data)} records to {output_file}")

        except Exception as e:
            logging.error(f"Error writing output: {e}")
            raise

    def _populate_qc_sheet(self, workbook):
        """Automatically populate QC sheet with data from QC file into existing columns"""
        try:
            import re
            from openpyxl.utils import get_column_letter
            
            # Check if QC sheet exists
            if "QC" not in workbook.sheetnames:
                logging.info("No existing QC sheet found, skipping QC data population")
                return
                
            qc_sheet = workbook["QC"]
            logging.info("Found existing QC sheet, populating with QC data")
            
            # Get complete row data from QC file
            qc_data_rows = self.qc_reader.get_qc_data_rows()
            
            if not qc_data_rows:
                logging.warning("No QC data rows found to populate")
                return
            
            logging.info(f"QC data rows to populate: {len(qc_data_rows)}")
            
            # Find the header row in the existing QC sheet (try rows 1, 2, 3)
            header_row = None
            existing_headers = {}
            
            for row_num in [1, 2, 3]:
                headers_found = {}
                for col_idx in range(1, min(qc_sheet.max_column + 1, 50)):  # Limit column scan to 50 columns
                    cell_value = qc_sheet.cell(row=row_num, column=col_idx).value
                    if cell_value:
                        # Clean header text (remove extra spaces, newlines)
                        header_text = re.sub(r"\s+", " ", str(cell_value).replace("\n", " ")).strip()
                        if header_text:
                            headers_found[header_text] = col_idx
                
                # Check if this row has the required columns
                if any(header in headers_found for header in ['Pole', 'To Pole']):
                    header_row = row_num
                    existing_headers = headers_found
                    logging.info(f"Found headers in row {header_row}: {list(existing_headers.keys())}")
                    break
            
            if not header_row:
                logging.warning("Could not find header row with 'Pole' and 'To Pole' columns in QC sheet")
                return
            
            # Clear existing data rows (keep headers intact) - but only clear what we need
            data_start_row = header_row + 1
            max_cols_to_clear = min(qc_sheet.max_column, 50)  # Limit to 50 columns
            
            # Find the actual end of data to avoid clearing unnecessary rows
            actual_max_row = data_start_row
            for row_idx in range(data_start_row, min(qc_sheet.max_row + 1, data_start_row + 10000)):  # Limit to 10,000 rows
                has_data = False
                for col_idx in range(1, max_cols_to_clear + 1):
                    if qc_sheet.cell(row=row_idx, column=col_idx).value:
                        has_data = True
                        break
                if has_data:
                    actual_max_row = row_idx
                else:
                    break  # Stop at first empty row
            
            # Only clear rows that actually have data
            rows_to_clear = min(actual_max_row - data_start_row + 1, len(qc_data_rows) + 100)  # Clear existing data + some buffer
            logging.info(f"Clearing {rows_to_clear} rows starting from row {data_start_row}")
            
            if rows_to_clear > 0:
                for row_idx in range(data_start_row, data_start_row + rows_to_clear):
                    for col_idx in range(1, max_cols_to_clear + 1):
                        qc_sheet.cell(row=row_idx, column=col_idx).value = None
            
            # Create mapping from QC file columns to existing sheet columns
            column_mapping = {}
            
            # Log available columns for debugging
            qc_columns = list(qc_data_rows[0].keys()) if qc_data_rows else []
            logging.info(f"QC file columns: {qc_columns}")
            logging.info(f"QC sheet headers: {list(existing_headers.keys())}")
            
            # Enhanced column mapping with fuzzy matching
            for qc_column in qc_columns:
                mapped = False
                
                # Try exact match first
                if qc_column in existing_headers:
                    column_mapping[qc_column] = existing_headers[qc_column]
                    mapped = True
                else:
                    # Try case-insensitive match
                    for existing_header, col_idx in existing_headers.items():
                        if qc_column.lower() == existing_header.lower():
                            column_mapping[qc_column] = col_idx
                            mapped = True
                            break
                
                # Try normalized matching (ignoring spaces, punctuation, newlines)
                if not mapped:
                    for existing_header, col_idx in existing_headers.items():
                        if self._columns_match(qc_column, existing_header):
                            column_mapping[qc_column] = col_idx
                            logging.info(f"Normalized match '{qc_column}' -> '{existing_header}'")
                            mapped = True
                            break
                
                if not mapped:
                    logging.warning(f"Could not map QC column '{qc_column}' to any QC sheet column")
            
            logging.info(f"Column mapping: {column_mapping}")
            logging.info(f"Mapped {len(column_mapping)} out of {len(qc_columns)} QC file columns")
            
            # Populate data rows
            rows_written = 0
            for row_idx, row_data in enumerate(qc_data_rows):
                sheet_row = data_start_row + row_idx
                
                # Populate mapped columns from QC file
                for qc_column, value in row_data.items():
                    if qc_column in column_mapping:
                        col_idx = column_mapping[qc_column]
                        qc_sheet.cell(row=sheet_row, column=col_idx).value = value
                
                # Try to populate missing columns from main data if available
                self._populate_missing_qc_columns(qc_sheet, sheet_row, row_data, existing_headers, column_mapping)
                
                rows_written += 1
                
                # Progress logging for large datasets
                if rows_written % 100 == 0:
                    logging.info(f"Populated {rows_written} QC rows...")
            
            logging.info(f"Successfully populated QC sheet with {rows_written} rows into {len(column_mapping)} matching columns")
            
        except Exception as e:
            logging.error(f"Error populating QC sheet: {e}")
            # Don't raise the exception - QC sheet population is optional
    
    def _columns_match(self, qc_column, sheet_column):
        """Check if QC file column matches QC sheet column by normalizing spaces, punctuation, and newlines"""
        import re
        
        def normalize_column_name(name):
            # Remove newlines and replace with spaces
            name = name.replace('\n', ' ').replace('\r', ' ')
            # Remove all punctuation and special characters, keep only letters, numbers, and spaces
            name = re.sub(r'[^\w\s]', ' ', name)
            # Replace multiple spaces with single space and strip
            name = re.sub(r'\s+', ' ', name).strip()
            return name.lower()
        
        qc_norm = normalize_column_name(qc_column)
        sheet_norm = normalize_column_name(sheet_column)
        
        # Direct match after normalization
        return qc_norm == sheet_norm
    
    def _populate_missing_qc_columns(self, qc_sheet, sheet_row, row_data, existing_headers, column_mapping):
        """Populate missing QC columns with data from main processing if available"""
        try:
            # Get pole and to_pole from current row
            pole = row_data.get('Pole', '').strip()
            to_pole = row_data.get('To Pole', '').strip()
            
            if not pole:
                return
            
            # Check if we have processed data for this pole
            if hasattr(self, '_processed_data_cache'):
                pole_data = self._processed_data_cache.get(pole)
                if pole_data:
                    # Populate Pole Address if missing
                    if 'Pole Address (if available)' in existing_headers and 'Pole Address (if available)' not in column_mapping:
                        address = pole_data.get('Pole Address', '')
                        if address:
                            col_idx = existing_headers['Pole Address (if available)']
                            qc_sheet.cell(row=sheet_row, column=col_idx).value = address
                    
                    # Populate Proposed height if missing
                    if 'Proposed height of new attachment point' in existing_headers and 'Proposed height of new attachment point' not in column_mapping:
                        proposed_height = pole_data.get('Proposed MetroNet', '')
                        if proposed_height:
                            col_idx = existing_headers['Proposed height of new attachment point']
                            qc_sheet.cell(row=sheet_row, column=col_idx).value = proposed_height
                    
                    # Populate other missing columns as needed
                    missing_mappings = {
                        'Secondary or Neutral Power Height (Height of Lowest Power Conductor or Equipment, excluding streetlights)': 'Power Height',
                        'Pole Height & Class': 'Pole Height & Class',
                        'Pole to Pole Span Length (from starting point)': 'Span Length',
                        'Final Mid Span Ground Clearance of Proposed Attachment': 'Proposed MetroNet_Midspan',
                        'Guy Size': 'Guy Size',
                        'Guy Lead': 'Guy Lead',
                        'Guy Direction': 'Guy Direction',
                        'Notes (Items that need to be performed by Consumers Energy or other Companies)': 'Notes'
                    }
                    
                    for qc_header, data_key in missing_mappings.items():
                        if qc_header in existing_headers and qc_header not in [list(existing_headers.keys())[col_idx-1] for col_idx in column_mapping.values()]:
                            value = pole_data.get(data_key, '')
                            if value:
                                col_idx = existing_headers[qc_header]
                                qc_sheet.cell(row=sheet_row, column=col_idx).value = value
        
        except Exception as e:
            logging.debug(f"Error populating missing QC columns for row {sheet_row}: {e}")
    
    def _get_internal_key(self, element, attribute):
        """Get internal key for mapping"""
        mappings = {
            "Pole": {
                "Number": "Pole",
                "SCID": "Pole",
                "Map": "Map",
                "Address": "Address", 
                "Height & Class": "Pole Height/Class",
                "Pole Height/Class": "Pole Height/Class",
                "MR Notes": "Notes",
                "To Pole": "To Pole",
                "Latitude": "Latitude",
                "Longitude": "Longitude",
                "Tag": "Pole Tag",
                "Number of Existing Risers": "Existing Risers",
                "Existing Risers": "Existing Risers",
                "Existing Structure Type": "Structure Type",
                "Existing Loading": "Existing Load",
                "Proposed Loading": "Proposed Load"
            },
            "New Guy": {
                "Size": "Guy Size",
                "Lead": "Guy Lead", 
                "Direction": "Guy Direction",
                "Required": "Guy Needed"
            },
            "Power": {
                "Lowest Height": "Power Height",
                "Height": "Power Height",
                "Lowest Midspan": "Power Midspan",
                "Midspan": "Power Midspan",
                "Lowest Type": "Power Type"
            },
            "Span": {
                "Length": "Span Length"
            },
            "System": {
                "Line Number": "Line No."
            },
            "Street Light": {
                "Lowest Height": "Street Light Height",
                "Height": "Street Light Height"
            },
            "Cable": {
                "Type1": "Cable Type 1",
                "Diameter1": "Cable Diameter 1",
                "Type2": "Cable Type 2", 
                "Diameter2": "Cable Diameter 2",
                "Total Bundle Diameter": "Total Bundle Diameter"
            },
            "Power Equipment": {
                "Equipment List": "Power Equipments"
            }
        }
        
        if element in mappings:
            return mappings[element].get(attribute)
        elif element in ["comm1", "comm2", "comm3", "comm4"]:
            if attribute == "Attachment Ht":
                return element
            elif attribute == "Midspan Ht":
                return f"{element}_Midspan"
        elif element in self.config["telecom_providers"]:
            if attribute == "Attachment Ht":
                return element
            elif attribute == "Midspan Ht":
                # Special case for Proposed MetroNet midspan
                if element == "Proposed MetroNet":
                    return "Proposed MetroNet_Midspan"
                else:
                    return f"{element}_Midspan"
        
        return None

    def generate_output_file(self, job_name, template_path):
        """Generate output file by copying template with job name, preserving file extension."""
        import shutil
        from pathlib import Path
        
        template = Path(template_path)
        if not template.exists():
            logging.error(f"Template file not found: {template_path}")
            return None
        
        # Preserve the original file extension (.xlsx or .xlsm)
        template_extension = template.suffix
        output_file = template.parent / f"{job_name} Spread Sheet{template_extension}"
        
        try:
            shutil.copy2(template, output_file)
            
            # Verify the copy was successful
            if not output_file.exists() or output_file.stat().st_size == 0:
                logging.error(f"Copied output file '{output_file}' is empty. Check the template file.")
                return None
            return output_file
        except Exception as e:
            logging.error(f"Error copying template file: {e}")
            return None
    
    def _write_data_to_worksheet(self, ws, sorted_data, mapping_data, sheet_name=None):
        """Write sorted_data to worksheet ws using mapping_data for column mapping."""
        import re
        # Get config settings
        header_row = self.config.get("output_settings", {}).get("header_row", 1)
        data_start_row = self.config.get("output_settings", {}).get("data_start_row", header_row + 2)
        
        # Check if this is the Alden sheet
        is_alden_sheet = sheet_name and sheet_name.strip().lower() == "alden"

        # Get headers from the worksheet
        headers = []
        for cell_obj in ws[header_row]:
            if cell_obj.value:
                header_text = re.sub(r"\s+", " ", str(cell_obj.value).replace("\n", " ")).strip()
                headers.append(header_text)
            else:
                headers.append("")
        col_map = {h: idx + 1 for idx, h in enumerate(headers) if h.strip()}

        # Build mapping from internal key to Excel column name
        internal_to_excel = {}
        for element, attribute, output_col_name in mapping_data:
            internal_key = self._get_internal_key(element, attribute)
            if internal_key and output_col_name.strip():
                internal_to_excel[internal_key] = output_col_name
                logging.debug(f"Mapping {element}:{attribute} -> {internal_key} -> {output_col_name}")
            else:
                logging.debug(f"Skipping mapping {element}:{attribute} -> {internal_key} -> {output_col_name}")
        
        successful_writes = 0
        missing_columns = set()
        
        for i, data_row_content in enumerate(sorted_data, start=1):
            # Use Excel row from template if available, otherwise use sequential
            excel_row = data_row_content.get('_excel_row', data_start_row + i - 1)
            data_row_content['Line No.'] = i
            
            for internal_name, value in data_row_content.items():
                # Skip internal Excel row field
                if internal_name == '_excel_row':
                    continue
                    
                excel_col_name = internal_to_excel.get(internal_name, "")
                if not excel_col_name:
                    continue
                
                # Skip Pole and ToPole columns - preserve template data, but allow Pole Tag
                if excel_col_name.lower() in ['pole', 'to pole']:
                    continue
                
                col = col_map.get(excel_col_name)
                if col:
                    try:
                        cell_to_write = ws.cell(row=excel_row, column=col)
                        
                        # Convert measurement values for Alden sheet
                        if is_alden_sheet and value:
                            converted_value = self._convert_measurement_for_alden(value)
                            if converted_value:
                                value = converted_value
                        
                        cell_to_write.value = value
                        successful_writes += 1
                        logging.debug(f"Wrote {internal_name}='{value}' to {excel_col_name} (col {col}, row {excel_row})")
                    except Exception as e:
                        logging.warning(f"Error writing cell: {e}")
                else:
                    missing_columns.add(excel_col_name)
                    logging.debug(f"Column not found: '{excel_col_name}' (internal: {internal_name}, available: {list(col_map.keys())})")
                    if False:  # Tension-related code removed
                        logging.warning(f"Column not found in worksheet: {excel_col_name} (internal: {internal_name})")
        if missing_columns:
            logging.info(f"Note: Some mapped columns not found in template: {', '.join(sorted(missing_columns))}")
        else:
            logging.info("All mapped columns found in template")
        logging.info(f"Successfully wrote {successful_writes} data cells (preserving Pole and ToPole from template)")
    
    def _write_data_simple(self, ws, sorted_data, sheet_name=None):
        """Write sorted_data to worksheet ws with no mapping (just as columns in order)."""
        # Note: Removed conditional formatting as requested
        
        # Check if this is the Alden sheet
        is_alden_sheet = sheet_name and sheet_name.strip().lower() == "alden"
        
        for row_idx, row_data in enumerate(sorted_data, start=1):
            # Check if this row represents a QC mismatch (for logging only)
            is_qc_mismatch = False
            if self.qc_reader and self.qc_reader.is_active():
                pole = row_data.get('Pole', '')
                to_pole = row_data.get('To Pole', '')
                if pole and to_pole:
                    is_qc_mismatch = not self.qc_reader.has_connection(pole, to_pole)
                    if is_qc_mismatch:
                        logging.debug(f"QC mismatch detected for row {row_idx}: {pole} -> {to_pole}")
            
            # Write data but skip Pole and ToPole columns to preserve template data
            col_idx = 1
            for key, value in row_data.items():
                if key.lower() not in ['pole', 'to pole']:
                    # Convert measurement values for Alden sheet
                    if is_alden_sheet and value:
                        converted_value = self._convert_measurement_for_alden(value)
                        if converted_value:
                            value = converted_value
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    col_idx += 1
                # Note: Removed conditional formatting/highlighting as requested

    def _process_qc_filtered_connections(self, connections_df, mappings, sections_df):
        """Process connections when QC file is active - use EXACT QC Pole and ToPole values in specified order"""
        result_data = []
        
        # Get ordered connections from QC file in ORIGINAL format
        qc_original_connections = self.qc_reader.get_original_ordered_connections()
        qc_normalized_connections = self.qc_reader.get_ordered_connections()
        
        logging.info(f"Processing {len(qc_original_connections)} QC connections in specified order")
        logging.info("QC Mode: Using EXACT original Pole and ToPole format from QC file")
        
        # Create lookup for connection data from Excel (bidirectional)
        connection_lookup = {}
        for _, conn in connections_df.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            if n1 in mappings['node_id_to_scid'] and n2 in mappings['node_id_to_scid']:
                scid1 = mappings['node_id_to_scid'][n1]
                scid2 = mappings['node_id_to_scid'][n2]
                
                conn_info = {
                    'connection_id': conn.get('connection_id', ''),
                    'span_distance': conn.get('span_distance', ''),
                    'node1_id': n1,
                    'node2_id': n2
                }
                
                # Store connection lookup (use sorted tuple as key to avoid duplication)
                connection_key = tuple(sorted([scid1, scid2]))
                connection_lookup[connection_key] = conn_info
        
        # Process QC connections in the exact order specified in QC file
        for i, (qc_pole_orig, qc_to_pole_orig) in enumerate(qc_original_connections):
            # Get the corresponding normalized versions for data lookup
            qc_pole_norm, qc_to_pole_norm = qc_normalized_connections[i]
            # Check if this connection exists in Excel data using normalized SCIDs
            connection_key = tuple(sorted([qc_pole_norm, qc_to_pole_norm]))
            conn_info = connection_lookup.get(connection_key)
            
            if not conn_info:
                logging.warning(f"QC connection {qc_pole_orig} -> {qc_to_pole_orig} not found in Excel data")
                # Always create a row for QC connections, even if no data is available
                pole_node_data = mappings['scid_to_row'].get(qc_pole_norm, {})
                to_pole_node_data = mappings['scid_to_row'].get(qc_to_pole_norm, {})
                
                # Try to find span distance from connections_df using different lookup approaches
                span_distance = ''
                
                # Try direct SCID lookup in connections_df
                for _, conn in connections_df.iterrows():
                    n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
                    
                    # Check if either node matches our SCIDs (direct or through mapping)
                    scid1 = mappings['node_id_to_scid'].get(n1, n1)
                    scid2 = mappings['node_id_to_scid'].get(n2, n2)
                    
                    # Check all possible combinations
                    if ((scid1 == qc_pole_norm and scid2 == qc_to_pole_norm) or
                        (scid1 == qc_to_pole_norm and scid2 == qc_pole_norm) or
                        (n1 == qc_pole_norm and n2 == qc_to_pole_norm) or
                        (n1 == qc_to_pole_norm and n2 == qc_pole_norm)):
                        span_distance = conn.get('span_distance', '')
                        if span_distance:
                            logging.info(f"Found span distance {span_distance} for QC connection {qc_pole_orig} -> {qc_to_pole_orig}")
                            break
                
                # Single data source: Only use exact SCID matching from 'connections' sheet
                # No fallback mechanisms - if no exact match found, span_distance remains empty
                
                # Create connection info with found span distance (or empty if not found)
                conn_info = {
                    'connection_id': '',
                    'span_distance': span_distance,
                    'node1_id': '',
                    'node2_id': ''
                }
            
            # Get node data for the pole specified in QC file (using normalized SCID for lookup)
            pole_node_data = mappings['scid_to_row'].get(qc_pole_norm, {})
            
            # Always create output row for QC connections, even if no data is available
            # This ensures both main sheet and QC sheet have exactly the same connections
            row_data = self._create_qc_output_row(
                qc_pole_orig,  # Use ORIGINAL format from QC file
                qc_to_pole_orig,  # Use ORIGINAL format from QC file
                qc_pole_norm,  # Pass normalized version for data lookup
                qc_to_pole_norm,  # Pass normalized version for data lookup
                conn_info, 
                pole_node_data, 
                mappings['scid_to_row'], 
                sections_df
            )
            
            if row_data:
                row_data = self._apply_end_marker(row_data)
                result_data.append(row_data)
                logging.debug(f"Added QC connection (exact original): {qc_pole_orig} -> {qc_to_pole_orig}")
            else:
                # If _create_qc_output_row fails, create a minimal row to ensure connection is included
                logging.info(f"Creating minimal row for QC connection: {qc_pole_orig} -> {qc_to_pole_orig}")
                minimal_row = {
                    'Pole': qc_pole_orig,
                    'To Pole': qc_to_pole_orig,
                    'Line No.': '',
                    'Span Length': conn_info.get('span_distance', ''),
                    'Pole Height & Class': '',
                    'Power Height': '',
                    'Streetlight (bottom of bracket)': '',
                    'Guy Size': '',
                    'Guy Lead': '',
                    'Guy Direction': '',
                    'Pole Address': '',
                    'Notes': 'QC connection - limited data available'
                }
                minimal_row = self._apply_end_marker(minimal_row)
                result_data.append(minimal_row)
                logging.debug(f"Added minimal QC connection: {qc_pole_orig} -> {qc_to_pole_orig}")
        
        logging.info(f"Generated {len(result_data)} QC-filtered output rows in exact QC order")
        return result_data
    
    def _create_qc_output_row(self, pole_orig, to_pole_orig, pole_norm, to_pole_norm, conn_info, pole_node_data, scid_to_row, sections_df):
        """Create output row for QC filtering using exact ORIGINAL QC Pole and ToPole values"""
        # Try to create a row using the standard method first
        row_data = self._create_output_row(pole_norm, to_pole_norm, conn_info, pole_node_data, scid_to_row, sections_df)
        
        if row_data:
            # Force the exact ORIGINAL QC values (override any logic that might change them)
            row_data['Pole'] = pole_orig
            row_data['To Pole'] = to_pole_orig
            
            # Apply span length tolerance logic if QC reader is available
            if self.qc_reader and self.qc_reader.is_active():
                logging.info(f"Checking span length tolerance for {pole_orig} -> {to_pole_orig}")
                qc_span = self.qc_reader.get_qc_span_length(pole_orig, to_pole_orig)
                logging.info(f"QC span for {pole_orig} -> {to_pole_orig}: '{qc_span}'")
                
                if qc_span:
                    excel_span = row_data.get('Span Length', '')
                    tolerance = self.config.get('processing_options', {}).get('span_length_tolerance', 3)
                    logging.info(f"Excel span: '{excel_span}', QC span: '{qc_span}', tolerance: {tolerance}")
                    
                    # Apply tolerance check and use QC span if within tolerance
                    final_span = self._apply_span_length_tolerance(excel_span, qc_span, tolerance)
                    row_data['Span Length'] = final_span
                    logging.info(f"Final span length for {pole_orig} -> {to_pole_orig}: '{final_span}'")
                else:
                    logging.info(f"No QC span length found for {pole_orig} -> {to_pole_orig}")
            
            logging.debug(f"QC Row: Pole={pole_orig}, To Pole={to_pole_orig} (original format preserved)")
        else:
            # If standard method fails, validate Pole/To Pole values before creating minimal row
            if not pole_orig or not to_pole_orig or pole_orig.strip() == '' or to_pole_orig.strip() == '':
                logging.warning(f"Skipping minimal QC row creation: invalid Pole/To Pole values - Pole='{pole_orig}', To Pole='{to_pole_orig}'")
                return None
            
            # Create a minimal row to ensure QC connection is included
            logging.debug(f"Creating minimal QC row for {pole_orig} -> {to_pole_orig}")
            row_data = {
                'Pole': pole_orig,
                'To Pole': to_pole_orig,
                'Line No.': '',
                'Span Length': self._format_span_distance(conn_info.get('span_distance', '')),
                'Pole Height & Class': '',
                'Power Height': '',
                'Streetlight (bottom of bracket)': '',
                'Guy Size': '',
                'Guy Lead': '',
                'Guy Direction': '',
                'Pole Address': '',
                'Notes': 'QC connection - limited data available',
                'Guy Needed': 'NO',  # Default for QC connections with limited data
                # Add empty values for communication fields
                'comm1': '',
                'comm2': '',
                'comm3': '',
                'comm4': '',
                'Proposed MetroNet': '',
                'Verizon': '',
                'AT&T': '',
                'Comcast': '',
                'Zayo': '',
                'Jackson ISD': '',
                'All_Comm_Heights': '',
                'Total_Comm_Count': '',
                'Power Midspan': '',
                'Street Light Height': '',
                'Existing Risers': '',
                'Map': ''
            }
        
        return row_data

    def _add_sheet_comparison_formatting(self, workbook, main_sheet_name):
        """Conditional formatting has been disabled as requested - logging comparison info instead"""
        try:
            # Check if both sheets exist
            if main_sheet_name not in workbook.sheetnames or "QC" not in workbook.sheetnames:
                logging.info("Cannot compare sheets - missing main sheet or QC sheet")
                return
            
            main_sheet = workbook[main_sheet_name]
            qc_sheet = workbook["QC"]
            
            # Get config settings for data range
            header_row = self.config.get("output_settings", {}).get("header_row", 3)
            data_start_row = self.config.get("output_settings", {}).get("data_start_row", 4)
            
            # Find the data range in main sheet
            max_row = main_sheet.max_row
            max_col = min(main_sheet.max_column, 50)  # Limit to 50 columns for performance
            
            if max_row < data_start_row:
                logging.info("No data rows found in main sheet for comparison")
                return
            
            logging.info(f"Sheet comparison available between {main_sheet_name} and QC sheets")
            logging.info(f"Data range: {max_col} columns from row {data_start_row} to {max_row}")
            logging.info("Note: Conditional formatting disabled - differences not highlighted")
            
        except Exception as e:
            logging.error(f"Error during sheet comparison check: {e}")

    def _apply_alden_qc_comparison(self, workbook):
        """Compare Alden sheet data with Alden QC file and apply conditional formatting"""
        try:
            import re
            from openpyxl.styles import PatternFill
            
            # Check if Alden sheet exists
            if "Alden" not in workbook.sheetnames:
                logging.info("No Alden sheet found, skipping Alden QC comparison")
                return
            
            alden_sheet = workbook["Alden"]
            logging.info("Found Alden sheet, performing QC comparison")
            
            # Find header row and columns
            pole_col = None
            mr_notes_col = None
            metro_attach_col = None
            metro_mid_col = None
            power_attach_col = None
            power_mid_col = None
            power_type_col = None
            street_light_col = None
            comm1_col = None
            comm2_col = None
            comm3_col = None
            comm1_mid_col = None
            comm2_mid_col = None
            comm3_mid_col = None
            detected_header_row = None
            
            logging.debug(f"Searching for headers in Alden sheet. Max column: {alden_sheet.max_column}")
            for row_num in [1, 2, 3]:
                headers_found = []
                for col_idx in range(1, min(alden_sheet.max_column + 1, 100)):
                    cell_value = alden_sheet.cell(row=row_num, column=col_idx).value
                    if cell_value:
                        header_text = re.sub(r"\s+", " ", str(cell_value).replace("\n", " ")).strip().lower()
                        headers_found.append((col_idx, header_text))
                        if 'pole' in header_text and col_idx == 1:  # Make sure it's the first column
                            pole_col = col_idx
                        if 'mr notes' in header_text or 'makereadynotes' in header_text:
                            mr_notes_col = col_idx
                        if 'metro attachment' in header_text or 'metro attach' in header_text:
                            metro_attach_col = col_idx
                        if 'metro mid' in header_text:
                            metro_mid_col = col_idx
                        # Check for "Lowest Power at Pole" first (more specific)
                        if 'lowest power at pole' in header_text and power_attach_col is None:
                            power_attach_col = col_idx
                        # Fallback to just "lowest power" if not already found and not "Lowest Power Type"
                        elif 'lowest power' in header_text and 'type' not in header_text and power_attach_col is None:
                            power_attach_col = col_idx
                        if 'lowest power at mid' in header_text:
                            power_mid_col = col_idx
                        if 'lowest power type' in header_text:
                            power_type_col = col_idx
                        if 'street light height' in header_text:
                            street_light_col = col_idx
                        if 'comm1' in header_text and 'mid' not in header_text:
                            comm1_col = col_idx
                        if 'comm2' in header_text and 'mid' not in header_text:
                            comm2_col = col_idx
                        if 'comm3' in header_text and 'mid' not in header_text:
                            comm3_col = col_idx
                        if 'comm1' in header_text and 'mid' in header_text:
                            comm1_mid_col = col_idx
                        if 'comm2' in header_text and 'mid' in header_text:
                            comm2_mid_col = col_idx
                        if 'comm3' in header_text and 'mid' in header_text:
                            comm3_mid_col = col_idx
                if pole_col and mr_notes_col:
                    detected_header_row = row_num
                    logging.info(f"Found columns at row {detected_header_row}: Pole={pole_col}, MR_Notes={mr_notes_col}")
                    if metro_attach_col:
                        logging.info(f"  Metro_Attachment={metro_attach_col}")
                    if metro_mid_col:
                        logging.info(f"  Metro_Mid={metro_mid_col}")
                    if power_attach_col:
                        logging.info(f"  Power_Attachment={power_attach_col} (Lowest Power at Pole)")
                    else:
                        logging.warning(f"  Power_Attachment column NOT FOUND - power height comparison will be skipped")
                    if power_mid_col:
                        logging.info(f"  Power_Mid={power_mid_col}")
                    if power_type_col:
                        logging.info(f"  Power_Type={power_type_col}")
                    logging.debug(f"  Comm1={comm1_col}, Comm2={comm2_col}, Comm3={comm3_col}, Comm1_Mid={comm1_mid_col}, Comm2_Mid={comm2_mid_col}, Comm3_Mid={comm3_mid_col}")
                    break
            
            if not pole_col or not mr_notes_col:
                logging.warning(f"Could not find Pole and MR Notes columns in Alden sheet. Pole_col: {pole_col}, MR_notes_col: {mr_notes_col}")
                return
            
            logging.info(f"Found Pole column: {pole_col}, MR Notes column: {mr_notes_col} in Alden sheet")
            
            # Find data range - use detected header row
            data_start = detected_header_row + 1 if detected_header_row else 2
            max_row = alden_sheet.max_row
            logging.info(f"Data range: rows {data_start} to {max_row}")
            
            # Define fill colors
            match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            not_found_fill = PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid")  # Light blue
            
            matches = 0
            mismatches = 0
            not_found = 0
            metro_attach_matches = 0
            metro_attach_mismatches = 0
            metro_mid_matches = 0
            metro_mid_mismatches = 0
            power_attach_matches = 0
            power_attach_mismatches = 0
            power_mid_matches = 0
            power_mid_mismatches = 0
            power_type_matches = 0
            power_type_mismatches = 0
            comm1_attach_matches = 0
            comm1_attach_mismatches = 0
            comm1_mid_matches = 0
            comm1_mid_mismatches = 0
            comm2_attach_matches = 0
            comm2_attach_mismatches = 0
            comm2_mid_matches = 0
            comm2_mid_mismatches = 0
            comm3_attach_matches = 0
            comm3_attach_mismatches = 0
            comm3_mid_matches = 0
            comm3_mid_mismatches = 0
            
            # Compare each data row
            logging.debug(f"Starting comparison. Data start row: {data_start}, Max row: {max_row}")
            for row_idx in range(data_start, max_row + 1):
                pole_value = alden_sheet.cell(row=row_idx, column=pole_col).value
                mr_notes_value = alden_sheet.cell(row=row_idx, column=mr_notes_col).value
                
                if not pole_value:
                    continue
                
                pole_str = str(pole_value).strip()
                
                # Log first few poles for debugging
                if row_idx <= data_start + 2:
                    logging.info(f"Row {row_idx}: Pole={pole_str}, MR_Notes='{str(mr_notes_value)[:50] if mr_notes_value else 'None'}...'")
                
                # Check if pole exists in QC file first
                if not self.alden_qc_reader.has_pole(pole_str):
                    # Pole not found in QC file - highlight Pole cell with light blue
                    not_found += 1
                    alden_sheet.cell(row=row_idx, column=pole_col).fill = not_found_fill
                    if row_idx <= data_start + 2:
                        logging.info(f"  Pole {pole_str} not found in Alden QC file - highlighted in light blue")
                    continue
                
                # Get MR notes from QC file (now we know the pole exists)
                qc_mr_notes = self.alden_qc_reader.get_mr_notes(pole_str)
                
                # Normalize MR notes for comparison - extract text after colon if exists
                template_notes_full = str(mr_notes_value).strip() if mr_notes_value else ""
                # Extract notes after colon from template
                if ':' in template_notes_full:
                    template_notes = template_notes_full.split(':', 1)[1].strip()
                else:
                    template_notes = template_notes_full
                
                qc_notes_norm = qc_mr_notes.strip()
                
                # Check if they match
                if template_notes == qc_notes_norm:
                    # Match - apply green fill
                    alden_sheet.cell(row=row_idx, column=mr_notes_col).fill = match_fill
                    matches += 1
                    logging.debug(f"Pole {pole_str}: Match")
                else:
                    # Mismatch - apply red fill
                    alden_sheet.cell(row=row_idx, column=mr_notes_col).fill = mismatch_fill
                    mismatches += 1
                    logging.debug(f"Pole {pole_str}: Mismatch (template: '{template_notes[:50]}...' vs QC: '{qc_notes_norm[:50]}...')")
                
                # Compare MetroNet heights if columns exist
                if metro_attach_col:
                    result = self._compare_metronet_attachment_height(
                        alden_sheet, row_idx, pole_str, metro_attach_col, match_fill, mismatch_fill)
                    if result:
                        metro_attach_matches += result.get('matches', 0)
                        metro_attach_mismatches += result.get('mismatches', 0)
                
                if metro_mid_col:
                    result = self._compare_metronet_midspan_height(
                        alden_sheet, row_idx, pole_str, metro_mid_col, match_fill, mismatch_fill)
                    if result:
                        metro_mid_matches += result.get('matches', 0)
                        metro_mid_mismatches += result.get('mismatches', 0)
                
                # Compare Power heights if columns exist
                if power_attach_col:
                    result = self._compare_power_attachment_height(
                        alden_sheet, row_idx, pole_str, power_attach_col, match_fill, mismatch_fill, street_light_col)
                    if result:
                        power_attach_matches += result.get('matches', 0)
                        power_attach_mismatches += result.get('mismatches', 0)
                elif row_idx == data_start:  # Log only for first row if column not found
                    logging.warning("'Lowest Power at Pole' column not detected in Alden sheet - power height comparison will be skipped")
                
                if power_mid_col:
                    result = self._compare_power_midspan_height(
                        alden_sheet, row_idx, pole_str, power_mid_col, match_fill, mismatch_fill)
                    if result:
                        power_mid_matches += result.get('matches', 0)
                        power_mid_mismatches += result.get('mismatches', 0)
                
                # Compare Power Type if column exists
                if power_type_col:
                    result = self._compare_power_type(
                        alden_sheet, row_idx, pole_str, power_type_col, match_fill, mismatch_fill)
                    if result:
                        power_type_matches += result.get('matches', 0)
                        power_type_mismatches += result.get('mismatches', 0)
                
                # Compare Communication heights if columns exist
                if comm1_col:
                    result = self._compare_comm_attachment_height(
                        alden_sheet, row_idx, pole_str, comm1_col, 1, match_fill, mismatch_fill)
                    if result:
                        comm1_attach_matches += result.get('matches', 0)
                        comm1_attach_mismatches += result.get('mismatches', 0)
                
                if comm1_mid_col:
                    result = self._compare_comm_midspan_height(
                        alden_sheet, row_idx, pole_str, comm1_mid_col, 1, match_fill, mismatch_fill)
                    if result:
                        comm1_mid_matches += result.get('matches', 0)
                        comm1_mid_mismatches += result.get('mismatches', 0)
                
                if comm2_col:
                    result = self._compare_comm_attachment_height(
                        alden_sheet, row_idx, pole_str, comm2_col, 2, match_fill, mismatch_fill)
                    if result:
                        comm2_attach_matches += result.get('matches', 0)
                        comm2_attach_mismatches += result.get('mismatches', 0)
                
                if comm2_mid_col:
                    result = self._compare_comm_midspan_height(
                        alden_sheet, row_idx, pole_str, comm2_mid_col, 2, match_fill, mismatch_fill)
                    if result:
                        comm2_mid_matches += result.get('matches', 0)
                        comm2_mid_mismatches += result.get('mismatches', 0)
                
                if comm3_col:
                    result = self._compare_comm_attachment_height(
                        alden_sheet, row_idx, pole_str, comm3_col, 3, match_fill, mismatch_fill)
                    if result:
                        comm3_attach_matches += result.get('matches', 0)
                        comm3_attach_mismatches += result.get('mismatches', 0)
                
                if comm3_mid_col:
                    result = self._compare_comm_midspan_height(
                        alden_sheet, row_idx, pole_str, comm3_mid_col, 3, match_fill, mismatch_fill)
                    if result:
                        comm3_mid_matches += result.get('matches', 0)
                        comm3_mid_mismatches += result.get('mismatches', 0)
            
            logging.info(f"Alden QC comparison complete: {matches} matches, {mismatches} mismatches, {not_found} poles not found")
            if metro_attach_col:
                logging.info(f"MetroNet Attachment heights: {metro_attach_matches} matches, {metro_attach_mismatches} mismatches")
            if metro_mid_col:
                logging.info(f"MetroNet Midspan heights: {metro_mid_matches} matches, {metro_mid_mismatches} mismatches")
            if power_attach_col:
                logging.info(f"Power Attachment heights: {power_attach_matches} matches, {power_attach_mismatches} mismatches")
            if power_mid_col:
                logging.info(f"Power Midspan heights: {power_mid_matches} matches, {power_mid_mismatches} mismatches")
            if power_type_col:
                logging.info(f"Power Type: {power_type_matches} matches, {power_type_mismatches} mismatches")
            if comm1_col:
                logging.info(f"Comm1 Attachment heights: {comm1_attach_matches} matches, {comm1_attach_mismatches} mismatches")
            if comm1_mid_col:
                logging.info(f"Comm1 Midspan heights: {comm1_mid_matches} matches, {comm1_mid_mismatches} mismatches")
            if comm2_col:
                logging.info(f"Comm2 Attachment heights: {comm2_attach_matches} matches, {comm2_attach_mismatches} mismatches")
            if comm2_mid_col:
                logging.info(f"Comm2 Midspan heights: {comm2_mid_matches} matches, {comm2_mid_mismatches} mismatches")
            if comm3_col:
                logging.info(f"Comm3 Attachment heights: {comm3_attach_matches} matches, {comm3_attach_mismatches} mismatches")
            if comm3_mid_col:
                logging.info(f"Comm3 Midspan heights: {comm3_mid_matches} matches, {comm3_mid_mismatches} mismatches")
            
        except Exception as e:
            logging.error(f"Error during Alden QC comparison: {e}")
            import traceback
            logging.error(traceback.format_exc())
    
    def _compare_metronet_attachment_height(self, alden_sheet, row_idx, pole_str, col_idx, match_fill, mismatch_fill):
        """
        Compare MetroNet attachment height from template with QC data
        
        Returns:
            dict: {'matches': count, 'mismatches': count} or None if no comparison done
        """
        try:
            from .utils import Utils
            
            template_value = alden_sheet.cell(row=row_idx, column=col_idx).value
            qc_value = self.alden_qc_reader.get_metronet_attachment_height(pole_str)
            
            if not qc_value:
                # No QC data available, skip
                return None
            
            if not template_value:
                # Template has no value but QC does - mismatch
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: MetroNet attachment height mismatch (template: empty vs QC: '{qc_value}')")
                return {'matches': 0, 'mismatches': 1}
            
            # Convert both to comparable format
            template_str = str(template_value).strip()
            qc_str = str(qc_value).strip()
            
            # If template is in Alden format (e.g., "22ft 1in"), convert QC to same format if needed
            # If template is in decimal, convert both to decimal
            # For now, do simple string comparison
            if template_str == qc_str:
                alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                logging.debug(f"Pole {pole_str}: MetroNet attachment height match '{template_str}'")
                return {'matches': 1, 'mismatches': 0}
            else:
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: MetroNet attachment height mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                return {'matches': 0, 'mismatches': 1}
            
        except Exception as e:
            logging.error(f"Error comparing MetroNet attachment height for pole {pole_str}: {e}")
            return None
    
    def _compare_metronet_midspan_height(self, alden_sheet, row_idx, pole_str, col_idx, match_fill, mismatch_fill):
        """
        Compare MetroNet midspan height from template with QC data
        
        Returns:
            dict: {'matches': count, 'mismatches': count} or None if no comparison done
        """
        try:
            from .utils import Utils
            
            template_value = alden_sheet.cell(row=row_idx, column=col_idx).value
            qc_value = self.alden_qc_reader.get_metronet_midspan_height(pole_str)
            
            if not qc_value:
                # No QC data available, skip
                return None
            
            if not template_value:
                # Template has no value but QC does - mismatch
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: MetroNet midspan height mismatch (template: empty vs QC: '{qc_value}')")
                return {'matches': 0, 'mismatches': 1}
            
            # Convert both to comparable format
            template_str = str(template_value).strip()
            qc_str = str(qc_value).strip()
            
            # For multiple midspan values (comma-separated), check if any match
            if ',' in qc_str:
                qc_values = [v.strip() for v in qc_str.split(',')]
                # Check if template value matches any of the QC values
                if template_str in qc_values:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                    logging.debug(f"Pole {pole_str}: MetroNet midspan height match '{template_str}' (QC: {qc_str})")
                    return {'matches': 1, 'mismatches': 0}
                else:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                    logging.debug(f"Pole {pole_str}: MetroNet midspan height mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                    return {'matches': 0, 'mismatches': 1}
            else:
                # Single value comparison
                if template_str == qc_str:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                    logging.debug(f"Pole {pole_str}: MetroNet midspan height match '{template_str}'")
                    return {'matches': 1, 'mismatches': 0}
                else:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                    logging.debug(f"Pole {pole_str}: MetroNet midspan height mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                    return {'matches': 0, 'mismatches': 1}
            
        except Exception as e:
            logging.error(f"Error comparing MetroNet midspan height for pole {pole_str}: {e}")
            return None
    
    def _compare_power_attachment_height(self, alden_sheet, row_idx, pole_str, col_idx, match_fill, mismatch_fill, street_light_col=None):
        """
        Compare Power attachment height from template with QC data
        Includes Street Light if it's lower than Power height for comparison
        
        Returns:
            dict: {'matches': count, 'mismatches': count} or None if no comparison done
        """
        try:
            from .utils import Utils
            
            template_value = alden_sheet.cell(row=row_idx, column=col_idx).value
            qc_value = self.alden_qc_reader.get_power_attachment_height(pole_str)
            
            if not qc_value:
                # No QC data available, skip but log it for debugging
                logging.debug(f"Pole {pole_str}: No QC power attachment height data available (skipping comparison)")
                return None
            
            # Check if Street Light exists and is lower than Power
            effective_template_value = template_value
            if street_light_col:
                street_light_value = alden_sheet.cell(row=row_idx, column=street_light_col).value
                if street_light_value:
                    try:
                        # Convert both heights to decimal for comparison
                        template_decimal = Utils.parse_height_decimal(str(template_value).strip()) if template_value else None
                        street_light_decimal = Utils.parse_height_decimal(str(street_light_value).strip())
                        
                        if template_decimal and street_light_decimal and street_light_decimal > 0:
                            # Use Street Light if it's lower
                            if street_light_decimal < template_decimal:
                                effective_template_value = street_light_value
                                logging.debug(f"Pole {pole_str}: Using Street Light height {street_light_value} instead of Power {template_value} for comparison")
                    except Exception as e:
                        logging.debug(f"Error comparing Street Light height: {e}")
            
            if not effective_template_value:
                # Template has no value but QC does - mismatch
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: Power attachment height mismatch (template: empty vs QC: '{qc_value}')")
                return {'matches': 0, 'mismatches': 1}
            
            # Convert both to comparable format
            template_str = str(effective_template_value).strip()
            qc_str = str(qc_value).strip()
            
            # If template is in Alden format (e.g., "22ft 1in"), convert QC to same format if needed
            # If template is in decimal, convert both to decimal
            # For now, do simple string comparison
            if template_str == qc_str:
                alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                logging.debug(f"Pole {pole_str}: Power attachment height match '{template_str}'")
                return {'matches': 1, 'mismatches': 0}
            else:
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: Power attachment height mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                return {'matches': 0, 'mismatches': 1}
            
        except Exception as e:
            logging.error(f"Error comparing Power attachment height for pole {pole_str}: {e}")
            return None
    
    def _compare_power_midspan_height(self, alden_sheet, row_idx, pole_str, col_idx, match_fill, mismatch_fill):
        """
        Compare Power midspan height from template with QC data
        
        Returns:
            dict: {'matches': count, 'mismatches': count} or None if no comparison done
        """
        try:
            from .utils import Utils
            
            template_value = alden_sheet.cell(row=row_idx, column=col_idx).value
            qc_value = self.alden_qc_reader.get_power_midspan_height(pole_str)
            
            if not qc_value:
                # No QC data available, skip
                return None
            
            if not template_value:
                # Template has no value but QC does - mismatch
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: Power midspan height mismatch (template: empty vs QC: '{qc_value}')")
                return {'matches': 0, 'mismatches': 1}
            
            # Convert both to comparable format
            template_str = str(template_value).strip()
            qc_str = str(qc_value).strip()
            
            # For multiple midspan values (comma-separated), check if any match
            if ',' in qc_str:
                qc_values = [v.strip() for v in qc_str.split(',')]
                # Check if template value matches any of the QC values
                if template_str in qc_values:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                    logging.debug(f"Pole {pole_str}: Power midspan height match '{template_str}' (QC: {qc_str})")
                    return {'matches': 1, 'mismatches': 0}
                else:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                    logging.debug(f"Pole {pole_str}: Power midspan height mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                    return {'matches': 0, 'mismatches': 1}
            else:
                # Single value comparison
                if template_str == qc_str:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                    logging.debug(f"Pole {pole_str}: Power midspan height match '{template_str}'")
                    return {'matches': 1, 'mismatches': 0}
                else:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                    logging.debug(f"Pole {pole_str}: Power midspan height mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                    return {'matches': 0, 'mismatches': 1}
            
        except Exception as e:
            logging.error(f"Error comparing Power midspan height for pole {pole_str}: {e}")
            return None
    
    def _compare_power_type(self, alden_sheet, row_idx, pole_str, col_idx, match_fill, mismatch_fill):
        """
        Compare Power Type from template with QC AttachmentType using substring matching
        
        Returns:
            dict: {'matches': count, 'mismatches': count} or None if no comparison done
        """
        try:
            template_value = alden_sheet.cell(row=row_idx, column=col_idx).value
            qc_attachment_type = self.alden_qc_reader.get_power_attachment_type(pole_str)
            
            if not qc_attachment_type:
                # No QC data available, skip
                return None
            
            if not template_value:
                # Template has no value but QC does - mismatch
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: Power Type mismatch (template: empty vs QC: '{qc_attachment_type}')")
                return {'matches': 0, 'mismatches': 1}
            
            template_str = str(template_value).strip()
            qc_str = str(qc_attachment_type).strip()
            
            # Case-insensitive substring matching with intelligent keyword extraction
            # Examples:
            # - "Secondary Drip Loop" (template) matches "Secondary" (QC) → True
            # - "Secondary" (template) matches "Power Secondary" (QC) → True
            # - "Secondary Drip Loop" (template) matches "Power Secondary" (QC) → True
            template_lower = template_str.lower()
            qc_lower = qc_str.lower()
            
            # First check: simple substring matching (either contains the other)
            is_match = (template_lower in qc_lower) or (qc_lower in template_lower)
            
            # Second check: extract base keywords and compare
            # This handles cases like "Secondary Drip Loop" vs "Power Secondary"
            # Both contain "Secondary" as a base keyword, so they should match
            if not is_match:
                import re
                
                # Get power keywords from config for base keyword extraction
                power_keywords = self.config.get("power_keywords", [])
                
                # Extract words from both strings
                template_words = set(re.findall(r'\b\w+\b', template_lower))
                qc_words = set(re.findall(r'\b\w+\b', qc_lower))
                
                # Check if any power keyword matches between template and QC
                # Normalize power keywords to lowercase for comparison
                normalized_power_keywords = {kw.strip().lower() for kw in power_keywords if kw.strip()}
                
                # Find matching power keywords in both template and QC
                template_keywords = template_words.intersection(normalized_power_keywords)
                qc_keywords = qc_words.intersection(normalized_power_keywords)
                
                # If they share any power keyword, it's a match
                if template_keywords and qc_keywords:
                    is_match = len(template_keywords.intersection(qc_keywords)) > 0
                
                # Fallback: word boundary matching
                if not is_match:
                    template_word = r'\b' + re.escape(template_lower) + r'\b'
                    qc_word = r'\b' + re.escape(qc_lower) + r'\b'
                    is_match = (re.search(template_word, qc_lower) is not None) or (re.search(qc_word, template_lower) is not None)
            
            if is_match:
                alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                logging.debug(f"Pole {pole_str}: Power Type match (template: '{template_str}' vs QC: '{qc_str}')")
                return {'matches': 1, 'mismatches': 0}
            else:
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: Power Type mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                return {'matches': 0, 'mismatches': 1}
                
        except Exception as e:
            logging.error(f"Error comparing Power Type for pole {pole_str}: {e}")
            return None
    
    def _compare_comm_attachment_height(self, alden_sheet, row_idx, pole_str, col_idx, comm_number, match_fill, mismatch_fill):
        """
        Compare communication attachment height from template with QC data
        
        Returns:
            dict: {'matches': count, 'mismatches': count} or None if no comparison done
        """
        try:
            from .utils import Utils
            
            template_value = alden_sheet.cell(row=row_idx, column=col_idx).value
            qc_value = self.alden_qc_reader.get_comm_attachment_height(pole_str, comm_number)
            
            if not qc_value:
                # No QC data available, skip
                return None
            
            if not template_value:
                # Template has no value but QC does - mismatch
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: Comm{comm_number} attachment height mismatch (template: empty vs QC: '{qc_value}')")
                return {'matches': 0, 'mismatches': 1}
            
            # Convert both to comparable format
            template_str = str(template_value).strip()
            qc_str = str(qc_value).strip()
            
            # If template is in Alden format (e.g., "22ft 1in"), convert QC to same format if needed
            # If template is in decimal, convert both to decimal
            # For now, do simple string comparison
            if template_str == qc_str:
                alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                logging.debug(f"Pole {pole_str}: Comm{comm_number} attachment height match '{template_str}'")
                return {'matches': 1, 'mismatches': 0}
            else:
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: Comm{comm_number} attachment height mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                return {'matches': 0, 'mismatches': 1}
            
        except Exception as e:
            logging.error(f"Error comparing Comm{comm_number} attachment height for pole {pole_str}: {e}")
            return None
    
    def _compare_comm_midspan_height(self, alden_sheet, row_idx, pole_str, col_idx, comm_number, match_fill, mismatch_fill):
        """
        Compare communication midspan height from template with QC data
        
        Returns:
            dict: {'matches': count, 'mismatches': count} or None if no comparison done
        """
        try:
            from .utils import Utils
            
            template_value = alden_sheet.cell(row=row_idx, column=col_idx).value
            qc_value = self.alden_qc_reader.get_comm_midspan_height(pole_str, comm_number)
            
            if not qc_value:
                # No QC data available, skip
                return None
            
            if not template_value:
                # Template has no value but QC does - mismatch
                alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                logging.debug(f"Pole {pole_str}: Comm{comm_number} midspan height mismatch (template: empty vs QC: '{qc_value}')")
                return {'matches': 0, 'mismatches': 1}
            
            # Convert both to comparable format
            template_str = str(template_value).strip()
            qc_str = str(qc_value).strip()
            
            # For multiple midspan values (comma-separated), check if any match
            if ',' in qc_str:
                qc_values = [v.strip() for v in qc_str.split(',')]
                # Check if template value matches any of the QC values
                if template_str in qc_values:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                    logging.debug(f"Pole {pole_str}: Comm{comm_number} midspan height match '{template_str}' (QC: {qc_str})")
                    return {'matches': 1, 'mismatches': 0}
                else:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                    logging.debug(f"Pole {pole_str}: Comm{comm_number} midspan height mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                    return {'matches': 0, 'mismatches': 1}
            else:
                # Single value comparison
                if template_str == qc_str:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = match_fill
                    logging.debug(f"Pole {pole_str}: Comm{comm_number} midspan height match '{template_str}'")
                    return {'matches': 1, 'mismatches': 0}
                else:
                    alden_sheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill
                    logging.debug(f"Pole {pole_str}: Comm{comm_number} midspan height mismatch (template: '{template_str}' vs QC: '{qc_str}')")
                    return {'matches': 0, 'mismatches': 1}
            
        except Exception as e:
            logging.error(f"Error comparing Comm{comm_number} midspan height for pole {pole_str}: {e}")
            return None

    def _apply_span_length_tolerance(self, excel_span, qc_span, tolerance):
        """
        Check if QC span length is within tolerance of Excel span length
        
        Args:
            excel_span (str): Span length from Excel data
            qc_span (str): Span length from QC data
            tolerance (float): Tolerance in feet
            
        Returns:
            str: QC span length if within tolerance, otherwise Excel span length
        """
        if not qc_span or not excel_span:
            return excel_span or qc_span
        
        try:
            # Convert both to numeric values - remove commas, single quotes, and extra spaces
            excel_clean = str(excel_span).replace(',', '').replace("'", '').strip()
            qc_clean = str(qc_span).replace(',', '').replace("'", '').strip()
            
            excel_value = float(excel_clean)
            qc_value = float(qc_clean)
            
            # Check if difference is within tolerance
            difference = abs(excel_value - qc_value)
            if difference <= tolerance:
                logging.info(f"Using QC span length {qc_span} (Excel: {excel_span}, difference: {difference:.1f}ft, tolerance: {tolerance}ft)")
                return qc_span
            else:
                logging.info(f"QC span length {qc_span} outside tolerance (Excel: {excel_span}, difference: {difference:.1f}ft, tolerance: {tolerance}ft) - using Excel value")
                return excel_span
                
        except (ValueError, TypeError) as e:
            logging.debug(f"Error comparing span lengths '{excel_span}' vs '{qc_span}': {e}")
            return excel_span or qc_span